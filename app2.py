from flask import Flask, render_template, request, jsonify, send_file, redirect, url_for, flash
import pandas as pd
import os
from datetime import datetime
import json
from werkzeug.utils import secure_filename
import tempfile
import zipfile
from io import BytesIO
import logging
import requests
from dotenv import load_dotenv
import re
from utils import AmountToWords, InvoiceNumberGenerator, AIIntegration, safe_float_conversion, safe_int_conversion, safe_date_conversion
from openpyxl import load_workbook   # 🔹 Added for Excel template handling

# Load environment variables
load_dotenv()

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)
app.secret_key = 'your-secret-key-here'
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Ensure upload folder exists
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

ALLOWED_EXTENSIONS = {'xlsx', 'xls', 'xlsm'}


# Initialize global objects
amount_converter = AmountToWords()
invoice_generator = InvoiceNumberGenerator()
ai_integration = AIIntegration()

# 🔹 Helper: Generate HLM Excel from Template
def generate_hlm_excel_from_template(bill, template_path="HLM_Template.xlsm", output_folder="hlm_bills"):
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    wb = load_workbook(template_path, keep_vba=True)   # ✅ VBA/macros को preserve करने के लिए

    # ================= Sheet1 (Invoice) =================
    ws1 = wb["Invoice"]
    ws1['B12'] = bill['centre_name']
    ws1['G11'] = bill['bill_date']
    ws1['G12'] = bill['bill_number']
    ws1['G13'] = datetime.strptime(bill['bill_date'], "%Y-%m-%d").strftime("%B-%Y")

    pathology_count = sum(1 for t in bill['test_items'] if t.get('test_type', '').lower() == "pathology")
    radiology_count = sum(1 for t in bill['test_items'] if t.get('test_type', '').lower() in ["radiology", "nuclear"])

    if pathology_count > 0:
        ws1['B19'] = "Pathology Investigation"
        ws1['E19'] = pathology_count
    if radiology_count > 0:
        ws1['B20'] = "Radiology Investigation"
        ws1['E20'] = radiology_count

    total_mrp = sum(t['mrp'] for t in bill['test_items'])
    total_sharing = sum(t['sharing_amount'] for t in bill['test_items'])
    ws1['G19'] = total_mrp
    ws1['G27'] = total_sharing

    # ================= Sheet2 (Detailed) =================
    ws2 = wb["Detailed"]

    # पुराना data clear (header row छोड़कर)
    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
        for cell in row:
            cell.value = None

    # नया data डालना
    for i, item in enumerate(bill['test_items'], 2):
        ws2[f"A{i}"] = item.get('registered_date', '')
        ws2[f"B{i}"] = item.get('patient_name', '')
        ws2[f"C{i}"] = item.get('visit_code', '')
        ws2[f"D{i}"] = item.get('test_name', '')
        ws2[f"E{i}"] = item.get('test_type', '')
        ws2[f"F{i}"] = item.get('mrp', 0)
        ws2[f"G{i}"] = item.get('sharing_amount', 0)
        ws2[f"H{i}"] = item.get('rate', 0)  # Net Amount = Rate

    safe_center_name = str(bill['centre_name']).replace(" ", "_").replace("/", "_")
    output_path = os.path.join(output_folder, f"{safe_center_name}.xlsm")
    wb.save(output_path)

    return output_path
    # ---------------- Existing Code Below ---------------- #

def allowed_file(filename):
    """Check if file extension is allowed"""
    if not filename or '.' not in filename:
        return False
    return filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def validate_excel_data(df):
    """Validate Excel data structure and required columns"""
    required_columns = ['RegisteredDate', 'PatientVisitCode', 'PatientName', 'TEST NAME', 'MRP', 'CentreTestRate', 'CENTER NAME']
    
    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        return False, f"Missing required columns: {', '.join(missing_columns)}"
    
    # Check for empty dataframe
    if df.empty:
        return False, "Excel file is empty"
    
    # Check for required data in key columns
    if df['CENTER NAME'].isna().all():
        return False, "No center names found in the data"
    
    if df['MRP'].isna().all():
        return False, "No MRP values found in the data"
    
    return True, None



def process_excel_file(file_path):
    """Process Excel file and return billing data grouped by Center Name, with MobileNumber filtering."""
    try:
        df = pd.read_excel(file_path)
        is_valid, error_msg = validate_excel_data(df)
        if not is_valid:
            return None, error_msg
        df = df.fillna('')
        # Store original df for later filtering
        return df, None
    except Exception as e:
        logger.error(f"Error processing Excel file: {e}")
        return None, f"Error processing file: {str(e)}"

def get_hlm_centers():
    """Get list of HLM centers"""
    hlm_centers = [
        'Manglam_Diagnostics_Agroha',
        'Hansi_Lab_MANGLAM',
        'JIND_MANGLAM_LAB_HISAR',
        'Narwana_Manglam_Lab',
        'Sanyam_Lab,_Inside_Satija_Healthcare,_H..',
        'Shri_Bala_JI_Lab_Isherwal,_Miran',
        'Vikash_Laboratory,_Java_Hospital,_Tosha..'
    ]
    return hlm_centers

def get_b2b_centers():
    """Get list of B2B centers (non-HLM)"""
    all_centers = [
        'Aarogya_Hospital_Hisar',
        'AMANDEEP_HOSPITAL',
        'AMARAVATI_HOSPITAL',
        'APEX_DIAGNOSTICS',
        'Barwala_Jansevarth_Lab,_Barwala',
        'CITY_CENTER_HISAR',
        'DR_ANKIT_GOYAL',
        'DR._RAJESH_MEHTA',
        'ECHS',
        'Elora_Dass_Gupta',
        'Fatehabad_Manglam_Diagnostices',
        'Geetanjali_Hospital',
        'GOBIND_NARSING_HOME',
        'GUPTA_NEWTON_HOSPITAL',
        'Guru_Jambheshwar_Multispeciality_Hosp....',
        'HISAR_DIAGNOSTICS_JHAJHPUL',
        'HISAR_GESTRO_HOSPITAL',
        'Hisar_Hospital_Nursery_Inside_Hsr_Hosp..',
        'HOLY_HELP_HOSPITAL',
        'INSURANCE,_HISAR',
        'JANKI_HOSPITAL',
        'LIFE_LINE_HOSPITAL',
        'MEYANSH_HOSPITAL',
        'Navjeevan_Hospital',
        'Onquest_Laboratories_Ltd..',
        'Pathkind_Diagnostics',
        'Ram_Niwas\'s_Centre',
        'Ravindra_Hospital',
        'RMCT_TOHANA',
        'SACHIN_MITTAL',
        'SADBHAVNA_HOSPITAL',
        'Sai_Hospital',
        'Sapra_Hospital,_Hisar',
        'SARVODYA_HOSPITAL',
        'SHANI_MANAV_SEVA_TRUST',
        'SHANTI_GI_HOSPITAL',
        'SHARDHA_HOSPITAL',
        'Shree_Krishna_Pranami_Multi_speciality_H..'
    ]
    hlm_centers = get_hlm_centers()
    b2b_centers = [center for center in all_centers if center not in hlm_centers]
    return b2b_centers

@app.route('/')
def index():
    try:
        return render_template('index.html', app=app)
    except Exception as e:
        logger.error(f"Error in index route: {e}")
        flash('An error occurred while loading the page', 'error')
    return render_template('index.html', app=app)

@app.route('/upload', methods=['POST'])
def upload_file():
    try:
        if 'file' not in request.files:
            flash('No file selected', 'error')
            return redirect(url_for('index'))
        file = request.files['file']
        if file.filename == '':
            flash('No file selected', 'error')
            return redirect(url_for('index'))
        if not file or not allowed_file(file.filename):
            flash('Invalid file type. Please upload an Excel file (.xlsx or .xls)', 'error')
            return redirect(url_for('index'))
        try:
            filename = secure_filename(file.filename)
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(file_path)
        except Exception as e:
            logger.error(f"Error saving file: {e}")
            flash('Error saving uploaded file', 'error')
            return redirect(url_for('index'))
        df, error = process_excel_file(file_path)
        if error:
            ai_suggestion = ai_integration.handle_error(error, "File upload and processing")
            flash(f'Error processing file: {error}\n\nAI Suggestion:\n{ai_suggestion}', 'error')
            return redirect(url_for('index'))
        if df is None or df.empty:
            flash('No bills could be generated from the uploaded file', 'error')
            return redirect(url_for('index'))
        app.df = df
        flash(f'Successfully uploaded {filename}', 'success')
        return redirect(url_for('index'))
    except Exception as e:
        logger.error(f"Error in upload_file: {e}")
        ai_suggestion = ai_integration.handle_error(str(e), "File upload process")
        flash(f'An unexpected error occurred while processing the file\n\nAI Suggestion:\n{ai_suggestion}', 'error')
        return redirect(url_for('index'))

@app.route('/generate_all_bills')
def generate_all_bills():
    """Generate bills for all centers"""
    try:
        if not hasattr(app, 'bills') or not app.bills:
            flash('No bills available. Please upload an Excel file first.', 'error')
            return redirect(url_for('index'))
        
        flash('All bills generated successfully!', 'success')
        return redirect(url_for('bills'))
    except Exception as e:
        logger.error(f"Error in generate_all_bills: {e}")
        flash('An error occurred while generating bills', 'error')
        return redirect(url_for('bills'))

@app.route('/generate_manual_bill', methods=['GET', 'POST'])
def generate_manual_bill():
    """Show manual bill generation page or process single bill generation"""
    try:
        if not hasattr(app, 'bills') or not app.bills:
            flash('No bills available. Please upload an Excel file first.', 'error')
            return redirect(url_for('index'))
        
        if request.method == 'POST':
            center_name = request.form.get('center_name')
            if not center_name:
                flash('Please select a center', 'error')
                return redirect(url_for('generate_manual_bill'))
            
            # Find the selected bill
            selected_bill = None
            for bill in app.bills:
                if bill['centre_name'] == center_name:
                    selected_bill = bill
                    break
            
            if selected_bill:
                # Create a new bills list with only the selected bill
                app.bills = [selected_bill]
                flash(f'Generated bill for {center_name}', 'success')
                return redirect(url_for('bills'))
            else:
                flash('Selected center not found', 'error')
                return redirect(url_for('generate_manual_bill'))
        
        return render_template('manual_bill.html', bills=app.bills, app=app)
    except Exception as e:
        logger.error(f"Error in generate_manual_bill: {e}")
        flash('An error occurred while processing manual bill generation', 'error')
        return redirect(url_for('bills'))

@app.route('/generate_multiple_bills', methods=['GET', 'POST'])
def generate_multiple_bills():
    """Show multiple bill generation page or process multiple bill generation"""
    try:
        if not hasattr(app, 'bills') or not app.bills:
            flash('No bills available. Please upload an Excel file first.', 'error')
            return redirect(url_for('index'))
        
        if request.method == 'POST':
            selected_centers = request.form.getlist('selected_centers')
            if not selected_centers:
                flash('Please select at least one center', 'error')
                return redirect(url_for('generate_multiple_bills'))
            
            # Filter bills for selected centers
            filtered_bills = [bill for bill in app.bills if bill['centre_name'] in selected_centers]
            
            if not filtered_bills:
                flash('No bills found for selected centers', 'error')
                return redirect(url_for('generate_multiple_bills'))
            
            app.bills = filtered_bills
            flash(f'Generated {len(filtered_bills)} bills for selected centers', 'success')
            return redirect(url_for('bills'))
        
        return render_template('multiple_bills.html', bills=app.bills, app=app)
    except Exception as e:
        logger.error(f"Error in generate_multiple_bills: {e}")
        flash('An error occurred while processing multiple bill generation', 'error')
        return redirect(url_for('bills'))

@app.route('/generate_hlm_bills', methods=['GET', 'POST'])
def generate_hlm_bills():
    """Two-step HLM bill generation: select center, then enter sharing per test type."""
    try:
        if not hasattr(app, 'df') or app.df.empty:
            flash('No data available. Please upload an Excel file first.', 'error')
            return redirect(url_for('index'))
        df = app.df.copy()
        # Step 1: Show HLM centers
        if request.method == 'GET' or (request.method == 'POST' and 'selected_center' not in request.form):
            hlm_df = df[df['MobileNumber'].astype(str).str.strip().str.upper() == 'HLM']
            centers = sorted(hlm_df['CENTER NAME'].dropna().unique())
            return render_template('hlm_bills.html', hlm_centers=centers, app=app)
        # Step 2: Center selected, show test types and sharing input
        selected_center = request.form.get('selected_center')
        if not selected_center:
            flash('Please select a center', 'error')
            return redirect(url_for('generate_hlm_bills'))
        center_df = df[(df['MobileNumber'].astype(str).str.strip().str.upper() == 'HLM') & (df['CENTER NAME'] == selected_center)]
        test_types = sorted(center_df['TEST TYPE'].dropna().unique())
        # If sharing percentages submitted, generate bill
        if 'submit_sharing' in request.form:
            sharing_map = {}
            for test_type in test_types:
                key = f"sharing_{test_type.replace(' ', '_').lower()}"
                val = request.form.get(key)
                try:
                    sharing_map[test_type] = float(val)
                except (TypeError, ValueError):
                    sharing_map[test_type] = 55.0
            # Generate bill for selected center
            test_items = []
            total_mrp = 0
            total_rate = 0
            total_sharing = 0
            for _, row in center_df.iterrows():
                mrp = safe_float_conversion(row['MRP'])
                test_type = str(row['TEST TYPE'])
                sharing_pct = sharing_map.get(test_type, 55.0)
                sharing_amount = mrp * (sharing_pct / 100)
                rate = mrp - sharing_amount
                test_item = {
                    'registered_date': safe_date_conversion(row['RegisteredDate']),
                    'visit_code': str(safe_int_conversion(row['PatientVisitCode'])),
                    'patient_name': str(row['PatientName']) if pd.notna(row['PatientName']) else 'N/A',
                    'test_name': str(row['TEST NAME']) if pd.notna(row['TEST NAME']) else 'N/A',
                    'test_type': test_type,
                    'mrp': mrp,
                    'rate': rate,
                    'sharing_amount': sharing_amount,
                    'sharing_percentage': sharing_pct
                }
                test_items.append(test_item)
                total_mrp += mrp
                total_rate += rate
                total_sharing += sharing_amount
            invoice_number = invoice_generator.generate()
            bill = {
                'centre_name': selected_center,
                'test_items': test_items,
                'test_types': test_types,
                'total_mrp': total_mrp,
                'total_rate': total_rate,
                'total_sharing': total_sharing,
                'bill_date': datetime.now().strftime('%Y-%m-%d'),
                'bill_number': invoice_number,
                'center_type': 'HLM',
                'amount_in_words': amount_converter.convert(total_rate)
            }
            app.bills = [bill]
              # 🔹 Excel Generation Call
            try:
                excel_path = generate_hlm_excel_from_template(bill)
                logger.info(f"HLM Excel generated: {excel_path}")
            except Exception as e:
                logger.error(f"Excel generation failed: {e}")

            flash(f'Generated HLM bill for {selected_center}', 'success')
            return redirect(url_for('bills'))

        return render_template('hlm_bills.html', selected_center=selected_center, test_types=test_types, app=app)
    except Exception as e:
        logger.error(f"Error in generate_hlm_bills: {e}")
        flash('An error occurred while processing HLM bill generation', 'error')
        return redirect(url_for('bills'))
          

@app.route('/generate_b2b_bills')
def generate_b2b_bills():
    """One-click B2B bill generation."""
    try:
        if not hasattr(app, 'df') or app.df.empty:
            flash('No data available. Please upload an Excel file first.', 'error')
            return redirect(url_for('index'))
        df = app.df.copy()
        b2b_df = df[df['MobileNumber'].astype(str).str.strip().str.upper() == 'B2B']
        bills = []
        for center_name, group in b2b_df.groupby('CENTER NAME'):
            if pd.isna(center_name) or center_name == '':
                continue
            test_items = []
            total_mrp = 0
            total_rate = 0
            total_sharing = 0
            for _, row in group.iterrows():
                mrp = safe_float_conversion(row['MRP'])
                rate = safe_float_conversion(row['CentreTestRate'])
                sharing_amount = mrp - rate
                test_item = {
                    'registered_date': safe_date_conversion(row['RegisteredDate']),
                    'visit_code': str(safe_int_conversion(row['PatientVisitCode'])),
                    'patient_name': str(row['PatientName']) if pd.notna(row['PatientName']) else 'N/A',
                    'test_name': str(row['TEST NAME']) if pd.notna(row['TEST NAME']) else 'N/A',
                    'test_type': str(row.get('TEST TYPE', 'Other')).strip(),
                    'mrp': mrp,
                    'rate': rate,
                    'sharing_amount': sharing_amount
                }
                test_items.append(test_item)
                total_mrp += mrp
                total_rate += rate
                total_sharing += sharing_amount
            invoice_number = invoice_generator.generate()
            bill = {
                'centre_name': str(center_name),
                'test_items': test_items,
                'total_mrp': total_mrp,
                'total_rate': total_rate,
                'total_sharing': total_sharing,
                'bill_date': datetime.now().strftime('%Y-%m-%d'),
                'bill_number': invoice_number,
                'center_type': 'B2B',
                'amount_in_words': amount_converter.convert(total_rate)
            }
            bills.append(bill)
        if not bills:
            flash('No B2B bills found in the uploaded data', 'error')
            return redirect(url_for('bills'))
        app.bills = bills
        flash(f'Generated {len(bills)} B2B bills', 'success')
        return redirect(url_for('bills'))
    except Exception as e:
        logger.error(f"Error in generate_b2b_bills: {e}")
        flash('An error occurred while processing B2B bill generation', 'error')
        return redirect(url_for('bills'))

@app.route('/bills')
def bills():
    try:
        if not hasattr(app, 'bills') or not app.bills:
            flash('No bills available. Please upload an Excel file first.', 'error')
            return redirect(url_for('index'))
        
        # Calculate total tests and amounts
        total_tests = sum(len(bill['test_items']) for bill in app.bills)
        total_mrp = sum(bill['total_mrp'] for bill in app.bills)
        total_rate = sum(bill['total_rate'] for bill in app.bills)
        total_sharing = sum(bill['total_sharing'] for bill in app.bills)
        
        return render_template('bills.html', 
                             bills=app.bills, 
                             total_tests=total_tests,
                             total_mrp=total_mrp,
                             total_rate=total_rate,
                             total_sharing=total_sharing,
                             app=app)
    except Exception as e:
        logger.error(f"Error in bills route: {e}")
        flash('An error occurred while loading bills', 'error')
        return redirect(url_for('index'))

@app.route('/bill/<int:bill_index>')
def view_bill(bill_index):
    try:
        if not hasattr(app, 'bills') or not app.bills:
            flash('No bills available', 'error')
            return redirect(url_for('bills'))
        
        if bill_index < 0 or bill_index >= len(app.bills):
            flash('Bill not found', 'error')
            return redirect(url_for('bills'))
        
        bill = app.bills[bill_index]
        return render_template('bill_detail.html', bill=bill, bill_index=bill_index, app=app)
    except Exception as e:
        logger.error(f"Error in view_bill: {e}")
        flash('An error occurred while viewing the bill', 'error')
        return redirect(url_for('bills'))

@app.route('/download_bill/<int:bill_index>')
def download_bill(bill_index):
    try:
        if not hasattr(app, 'bills') or not app.bills:
            flash('No bills available', 'error')
            return redirect(url_for('bills'))
        if bill_index < 0 or bill_index >= len(app.bills):
            flash('Bill not found', 'error')
            return redirect(url_for('bills'))
        bill = app.bills[bill_index]
        fmt = request.args.get('format', 'html').lower()
        if fmt == 'excel':
            # Excel in-memory using pandas
            buffer = BytesIO()
            df = pd.DataFrame(bill['test_items'])
            df.to_excel(buffer, index=False)
            buffer.seek(0)
            safe_center_name = str(bill['centre_name']).replace(" ", "_").replace("/", "_")
            filename = f"{safe_center_name}.xlsx"
            return send_file(buffer, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=filename)
        elif fmt == 'pdf':
            # PDF using pdfkit or xhtml2pdf
            html_content = render_template('bill_pdf.html', bill=bill)
            pdf_buffer = BytesIO()
            try:
                import pdfkit
                pdf = pdfkit.from_string(html_content, False)
                pdf_buffer.write(pdf)
                pdf_buffer.seek(0)
            except Exception:
                try:
                    from xhtml2pdf import pisa
                    pisa_status = pisa.CreatePDF(html_content, dest=pdf_buffer)
                    if pisa_status.err:
                        raise Exception("xhtml2pdf generation failed")
                    pdf_buffer.seek(0)
                except Exception as e:
                    logger.error(f"PDF generation failed: {e}")
                    flash('Could not generate PDF file for this bill', 'error')
                    return redirect(url_for('view_bill', bill_index=bill_index))
            safe_center_name = str(bill['centre_name']).replace(" ", "_").replace("/", "_")
            filename = f"{safe_center_name}.pdf"
            return send_file(pdf_buffer, mimetype='application/pdf', as_attachment=True, download_name=filename)
        else:
            html_content = render_template('bill_pdf.html', bill=bill)
            with tempfile.NamedTemporaryFile(mode='w', suffix='.html', delete=False, encoding='utf-8') as f:
                f.write(html_content)
                temp_path = f.name
            return send_file(temp_path, as_attachment=True, download_name=f"{bill['bill_number']}.html")
    except Exception as e:
        logger.error(f"Error in download_bill: {e}")
        flash('An error occurred while downloading the bill', 'error')
    return redirect(url_for('view_bill', bill_index=bill_index))

@app.route('/download_all_bills')
def download_all_bills():
    try:
        if not hasattr(app, 'bills') or not app.bills:
            flash('No bills available', 'error')
            return redirect(url_for('index'))
        
        # Create a ZIP file containing all bills
        memory_file = BytesIO()
        
        with zipfile.ZipFile(memory_file, 'w') as zf:
            for i, bill in enumerate(app.bills):
                try:
                    html_content = render_template('bill_pdf.html', bill=bill)
                    zf.writestr(f"{bill['bill_number']}.html", html_content)
                except Exception as e:
                    logger.error(f"Error processing bill {i}: {e}")
                    continue
        
        memory_file.seek(0)
        
        return send_file(
            memory_file,
            mimetype='application/zip',
            as_attachment=True,
            download_name=f"all_bills_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
        )
    except Exception as e:
        logger.error(f"Error in download_all_bills: {e}")
        flash('An error occurred while downloading all bills', 'error')
        return redirect(url_for('bills'))

@app.route('/download_all_excel')
def download_all_excel():
    try:
        if not hasattr(app, 'bills') or not app.bills:
            flash('No bills available', 'error')
            return redirect(url_for('bills'))
        memory_file = BytesIO()
        with zipfile.ZipFile(memory_file, 'w') as zf:
            for bill in app.bills:
                filename = f"{bill['bill_number']}.xlsx"
                filepath = os.path.join('bills', 'excel', filename)
                if os.path.exists(filepath):
                    zf.write(filepath, arcname=filename)
        memory_file.seek(0)
        return send_file(
            memory_file,
            mimetype='application/zip',
            as_attachment=True,
            download_name=f"all_bills_excel_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
        )
    except Exception as e:
        logger.error(f"Error in download_all_excel: {e}")
        flash('An error occurred while downloading all Excel bills', 'error')
        return redirect(url_for('bills'))

@app.route('/download_all_pdf')
def download_all_pdf():
    try:
        if not hasattr(app, 'bills') or not app.bills:
            flash('No bills available', 'error')
            return redirect(url_for('bills'))
        memory_file = BytesIO()
        with zipfile.ZipFile(memory_file, 'w') as zf:
            for bill in app.bills:
                filename = f"{bill['bill_number']}.pdf"
                filepath = os.path.join('bills', 'pdf', filename)
                if os.path.exists(filepath):
                    zf.write(filepath, arcname=filename)
        memory_file.seek(0)
        return send_file(
            memory_file,
            mimetype='application/zip',
            as_attachment=True,
            download_name=f"all_bills_pdf_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
        )
    except Exception as e:
        logger.error(f"Error in download_all_pdf: {e}")
        flash('An error occurred while downloading all PDF bills', 'error')
        return redirect(url_for('bills'))

@app.route('/api/bills')
def api_bills():
    try:
        if not hasattr(app, 'bills') or not app.bills:
            return jsonify({'error': 'No bills available'}), 404
        
        return jsonify(app.bills)
    except Exception as e:
        logger.error(f"Error in api_bills: {e}")
        return jsonify({'error': 'Internal server error'}), 500

@app.route('/ai_assistance', methods=['GET', 'POST'])
def ai_assistance():
    """AI assistance page for user queries"""
    if request.method == 'POST':
        user_query = request.form.get('user_query', '')
        if user_query:
            try:
                ai_response = ai_integration.get_ai_response(user_query)
                return jsonify({'response': ai_response})
            except Exception as e:
                logger.error(f"AI assistance error: {e}")
                return jsonify({'response': f"AI assistance temporarily unavailable: {str(e)}"})
    
    return render_template('ai_assistance.html', app=app)

# Error handlers
@app.errorhandler(404)
def not_found_error(error):
    return render_template('404.html'), 404

@app.errorhandler(500)
def internal_error(error):
    logger.error(f"Internal server error: {error}")
    return render_template('500.html'), 500

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000) 