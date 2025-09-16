#!/usr/bin/env python3
"""
Medical Test Billing Automation System
=====================================

A professional Python application for automating medical test billing for diagnostic centers.
Supports both B2B and HLM center types with configurable client sharing percentages.

Features:
- Excel and PDF bill generation
- AI-powered error handling and assistance
- Configurable billing parameters
- Professional invoice formatting
- Amount to words conversion
- Statutory compliance notes

Author: Medical Billing System
Version: 1.0.0
"""

import os
import sys
import argparse
import json
import logging
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Any
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors
import requests
from dotenv import load_dotenv
import re

# Load environment variables
load_dotenv()

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('billing.log'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

class AmountToWords:
    """Convert numerical amounts to words for invoice generation."""
    
    def __init__(self):
        self.units = ["", "One", "Two", "Three", "Four", "Five", "Six", "Seven", "Eight", "Nine"]
        self.teens = ["Ten", "Eleven", "Twelve", "Thirteen", "Fourteen", "Fifteen", "Sixteen", "Seventeen", "Eighteen", "Nineteen"]
        self.tens = ["", "", "Twenty", "Thirty", "Forty", "Fifty", "Sixty", "Seventy", "Eighty", "Ninety"]
        self.scales = ["", "Thousand", "Lakh", "Crore"]
    
    def convert(self, amount: float) -> str:
        """Convert amount to words in Indian numbering system."""
        if amount == 0:
            return "Zero Rupees Only"
        
        # Split into rupees and paise
        rupees = int(amount)
        paise = int(round((amount - rupees) * 100))
        
        rupees_words = self._convert_rupees(rupees)
        paise_words = self._convert_paise(paise) if paise > 0 else ""
        
        if paise_words:
            return f"{rupees_words} and {paise_words} Only"
        else:
            return f"{rupees_words} Only"
    
    def _convert_rupees(self, rupees: int) -> str:
        """Convert rupees to words."""
        if rupees == 0:
            return "Zero Rupees"
        
        words = []
        scale_index = 0
        
        while rupees > 0:
            chunk = rupees % 1000
            if chunk != 0:
                chunk_words = self._convert_chunk(chunk)
                if scale_index > 0:
                    chunk_words += f" {self.scales[scale_index]}"
                words.insert(0, chunk_words)
            
            rupees //= 1000
            scale_index += 1
        
        return " ".join(words) + " Rupees"
    
    def _convert_chunk(self, chunk: int) -> str:
        """Convert a chunk of up to 3 digits to words."""
        if chunk == 0:
            return ""
        
        words = []
        
        # Handle hundreds
        if chunk >= 100:
            words.append(f"{self.units[chunk // 100]} Hundred")
            chunk %= 100
        
        # Handle tens and units
        if chunk >= 20:
            words.append(self.tens[chunk // 10])
            if chunk % 10 > 0:
                words.append(self.units[chunk % 10])
        elif chunk >= 10:
            words.append(self.teens[chunk - 10])
        elif chunk > 0:
            words.append(self.units[chunk])
        
        return " ".join(words)
    
    def _convert_paise(self, paise: int) -> str:
        """Convert paise to words."""
        if paise == 0:
            return ""
        
        if paise < 20:
            return f"{self.units[paise]} Paise"
        else:
            tens = paise // 10
            units = paise % 10
            if units > 0:
                return f"{self.tens[tens]} {self.units[units]} Paise"
            else:
                return f"{self.tens[tens]} Paise"

class InvoiceNumberGenerator:
    """Generate sequential invoice numbers in the format KRPL/YY-YY/MM/NNN."""
    
    def __init__(self, start_sequence: int = 1):
        self.sequence = start_sequence
        self.current_year = datetime.now().year
        self.current_month = datetime.now().month
    
    def generate(self, invoice_date: datetime = None) -> str:
        """Generate invoice number for the given date."""
        if invoice_date is None:
            invoice_date = datetime.now()
        
        year = invoice_date.year
        month = invoice_date.month
        
        # Check if year or month has changed, reset sequence if needed
        if year != self.current_year or month != self.current_month:
            self.sequence = 1
            self.current_year = year
            self.current_month = month
        
        # Format: KRPL/YY-YY/MM/NNN
        year_range = f"{year-1}-{year}" if month < 4 else f"{year}-{year+1}"
        month_str = f"{month:02d}"
        sequence_str = f"{self.sequence:03d}"
        
        invoice_number = f"KRPL/{year_range}/{month_str}/{sequence_str}"
        
        # Increment sequence for next invoice
        self.sequence += 1
        
        # Validate sequence doesn't exceed 999
        if self.sequence > 999:
            logger.warning("Invoice sequence exceeded 999, resetting to 1")
            self.sequence = 1
        
        return invoice_number

class AIIntegration:
    """AI integration for error handling and user assistance."""
    
    def __init__(self):
        # Hardcode the Gemini API key directly
        self.gemini_api_key = "AIzaSyCISRlocKiVnAlakm5GEllJu6VVnrBdP6s"
        self.openai_api_key = None  # Disable OpenAI
        self.gemini_url = "https://generativelanguage.googleapis.com/v1/models/gemini-1.5-flash:generateContent"
    
    def get_ai_response(self, prompt: str, use_gemini: bool = True) -> str:
        """Get AI response for error handling or user assistance."""
        try:
            if use_gemini and self.gemini_api_key:
                return self._call_gemini(prompt)
            else:
                return "AI assistance not available. Please check your API keys."
        except Exception as e:
            logger.error(f"AI API call failed: {e}")
            return f"AI assistance temporarily unavailable: {str(e)}"
    
    def _call_gemini(self, prompt: str) -> str:
        """Call Gemini API via REST."""
        headers = {
            "Content-Type": "application/json",
        }
        
        data = {
            "contents": [{
                "parts": [{"text": prompt}]
            }]
        }
        
        response = requests.post(
            f"{self.gemini_url}?key={self.gemini_api_key}",
            headers=headers,
            json=data,
            timeout=30
        )
        
        if response.status_code == 200:
            result = response.json()
            return result['candidates'][0]['content']['parts'][0]['text']
        else:
            raise Exception(f"Gemini API error: {response.status_code}")
    
    def _call_openai(self, prompt: str) -> str:
        """Call OpenAI API."""
        try:
            import openai
            openai.api_key = self.openai_api_key
            
            response = openai.ChatCompletion.create(
                model="gpt-3.5-turbo",
                messages=[{"role": "user", "content": prompt}],
                max_tokens=500,
                timeout=30
            )
            
            return response.choices[0].message.content
        except ImportError:
            raise Exception("OpenAI library not installed. Run: pip install openai")
    
    def handle_error(self, error_message: str, context: str = "") -> str:
        """Get AI suggestions for error handling."""
        prompt = f"""
        I'm working with a medical billing automation system and encountered an error:
        
        Error: {error_message}
        Context: {context}
        
        Please provide:
        1. A brief explanation of what might be causing this error
        2. Step-by-step troubleshooting suggestions
        3. Any preventive measures to avoid this error in the future
        
        Keep the response concise and practical.
        """
        
        return self.get_ai_response(prompt)

class MedicalBillingProcessor:
    """Main billing processor for medical test billing automation."""
    
    def __init__(self, config: Dict[str, Any] = None):
        self.config = config or {}
        self.amount_converter = AmountToWords()
        self.invoice_generator = InvoiceNumberGenerator(
            self.config.get('invoice_sequence_start', 1)
        )
        self.ai = AIIntegration()
        
        # Create output directories
        self.output_dirs = {
            'excel': Path('bills/excel'),
            'pdf': Path('bills/pdf')
        }
        
        for dir_path in self.output_dirs.values():
            dir_path.mkdir(parents=True, exist_ok=True)
    
    def process_billing_data(self, main_data_path: str, supporting_data_path: str = None) -> List[Dict]:
        """Process main and supporting data to generate billing information."""
        try:
            # Load main data
            logger.info(f"Loading main data from {main_data_path}")
            main_df = pd.read_excel(main_data_path)
            
            # Load supporting data if provided
            supporting_df = None
            if supporting_data_path and os.path.exists(supporting_data_path):
                logger.info(f"Loading supporting data from {supporting_data_path}")
                supporting_df = pd.read_excel(supporting_data_path)
            
            # Clean and validate data
            main_df = self._clean_data(main_df)
            
            # Process billing by center
            bills = []
            for center_name, center_data in main_df.groupby('CENTER NAME'):
                if pd.isna(center_name) or center_name == '':
                    continue
                
                logger.info(f"Processing billing for center: {center_name}")
                bill_data = self._process_center_billing(
                    center_name, center_data, supporting_df
                )
                bills.append(bill_data)
            
            logger.info(f"Successfully processed {len(bills)} center bills")
            return bills
            
        except Exception as e:
            error_msg = f"Error processing billing data: {str(e)}"
            logger.error(error_msg)
            
            # Get AI assistance for error handling
            ai_suggestion = self.ai.handle_error(str(e), "Data processing stage")
            logger.info(f"AI suggestion: {ai_suggestion}")
            
            raise Exception(f"{error_msg}\n\nAI Suggestion:\n{ai_suggestion}")
    
    def _clean_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """Clean and validate the input data."""
        # Fill NaN values
        df = df.fillna('')
        
        # Ensure required columns exist
        required_columns = [
            'PatientVisitCode', 'RegisteredDate', 'PatientName', 'Age', 
            'AgeUnit', 'Gender', 'MobileNumber', 'TEST NAME', 'CODE NO', 
            'CENTER NAME', 'Modality', 'MRP', 'CentreTestRate', 'TEST TYPE'
        ]
        
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            raise ValueError(f"Missing required columns: {missing_columns}")
        
        # Convert date columns
        if 'RegisteredDate' in df.columns:
            df['RegisteredDate'] = pd.to_datetime(df['RegisteredDate'], errors='coerce')
        
        # Convert numeric columns
        numeric_columns = ['MRP', 'CentreTestRate', 'Age']
        for col in numeric_columns:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
        
        return df
    
    def _process_center_billing(self, center_name: str, center_data: pd.DataFrame, 
                               supporting_df: pd.DataFrame = None) -> Dict:
        """Process billing for a specific center."""
        # Determine center type from MobileNumber column
        center_type = self._determine_center_type(center_data)
        
        # Group by test type and calculate amounts
        test_summary = []
        total_gross = 0
        total_net = 0
        
        for test_type, test_group in center_data.groupby('TEST TYPE'):
            if pd.isna(test_type) or test_type == '':
                continue
            
            count = len(test_group)
            gross_amount = test_group['CentreTestRate'].sum()
            
            # Calculate net amount based on center type
            if center_type == 'B2B':
                net_amount = gross_amount
            else:  # HLM
                client_share_percentage = self._get_client_share_percentage(
                    center_name, test_type, supporting_df
                )
                net_amount = gross_amount * (client_share_percentage / 100)
            
            test_summary.append({
                'test_type': test_type,
                'count': count,
                'gross_amount': gross_amount,
                'net_amount': net_amount,
                'rate_per_test': gross_amount / count if count > 0 else 0.0
            })
            
            total_gross += gross_amount
            total_net += net_amount
        
        # Prepare patient details
        patient_details = []
        for idx, row in center_data.iterrows():
            try:
                patient_details.append({
                    'sr_no': idx + 1,
                    'date': row['RegisteredDate'].strftime('%d/%m/%Y') if pd.notna(row['RegisteredDate']) else 'N/A',
                    'patient_name': str(row['PatientName']) if pd.notna(row['PatientName']) else 'N/A',
                    'age': f"{row['Age']} {row['AgeUnit']}" if pd.notna(row['Age']) and pd.notna(row['AgeUnit']) else 'N/A',
                    'gender': str(row['Gender']) if pd.notna(row['Gender']) else 'N/A',
                    'test_name': str(row['TEST NAME']) if pd.notna(row['TEST NAME']) else 'N/A',
                    'value': float(row['CentreTestRate']) if pd.notna(row['CentreTestRate']) else 0.0
                })
            except Exception as e:
                logger.warning(f"Error processing patient row {idx}: {e}")
                # Add a placeholder entry to maintain sequence
                patient_details.append({
                    'sr_no': idx + 1,
                    'date': 'N/A',
                    'patient_name': 'Error in data',
                    'age': 'N/A',
                    'gender': 'N/A',
                    'test_name': 'N/A',
                    'value': 0.0
                })
        
        return {
            'center_name': center_name,
            'center_type': center_type,
            'test_summary': test_summary,
            'patient_details': patient_details,
            'total_gross': total_gross,
            'total_net': total_net,
            'invoice_date': self.config.get('invoice_date', datetime.now()),
            'period_start': self.config.get('period_start'),
            'period_end': self.config.get('period_end')
        }
    
    def _determine_center_type(self, center_data: pd.DataFrame) -> str:
        """Determine center type from MobileNumber column (expects exact 'B2B' or 'HLM')."""
        if 'MobileNumber' in center_data.columns:
            values = center_data['MobileNumber'].astype(str).str.strip().str.upper()
            if (values == 'B2B').any():
                return 'B2B'
            if (values == 'HLM').any():
                return 'HLM'
        return 'HLM'  # Default to HLM if not found
    
    def _get_client_share_percentage(self, center_name: str, test_type: str, 
                                    supporting_df: pd.DataFrame) -> float:
        """Get client sharing percentage from supporting data."""
        if supporting_df is None:
            return 50.0  # Default 50%
        
        try:
            # Validate supporting data has required columns
            required_columns = ['CENTER NAME', 'TEST TYPE', 'SHARE_PERCENTAGE']
            missing_columns = [col for col in required_columns if col not in supporting_df.columns]
            if missing_columns:
                logger.warning(f"Supporting data missing columns: {missing_columns}. Using default 50%")
                return 50.0
            
            # Look for matching center name and test type
            match = supporting_df[
                (supporting_df['CENTER NAME'].str.contains(center_name, case=False, na=False)) &
                (supporting_df['TEST TYPE'].str.contains(test_type, case=False, na=False))
            ]
            
            if not match.empty:
                # Return the first matching percentage
                share_percentage = match.iloc[0]['SHARE_PERCENTAGE']
                # Validate percentage is numeric and within valid range
                try:
                    share_percentage = float(share_percentage)
                    if 0 <= share_percentage <= 100:
                        return share_percentage
                    else:
                        logger.warning(f"Invalid share percentage {share_percentage}%. Using default 50%")
                        return 50.0
                except (ValueError, TypeError):
                    logger.warning(f"Invalid share percentage value. Using default 50%")
                    return 50.0
            else:
                return 50.0  # Default 50%
                
        except Exception as e:
            logger.warning(f"Error getting client share percentage: {e}")
            return 50.0  # Default 50%
    
    def generate_excel_bill(self, bill_data: Dict) -> str:
        """Generate Excel bill with summary and detailed sheets."""
        try:
            # Create workbook
            wb = openpyxl.Workbook()
            
            # Generate summary sheet
            self._create_summary_sheet(wb, bill_data)
            
            # Generate detailed sheet
            self._create_detailed_sheet(wb, bill_data)
            
            # Save file
            filename = f"{bill_data['center_name'].replace(' ', '_').replace('/', '_').replace('\\', '_')}.xlsx"
            filepath = self.output_dirs['excel'] / filename
            
            # Ensure output directory exists
            self.output_dirs['excel'].mkdir(parents=True, exist_ok=True)
            
            try:
                wb.save(filepath)
            except PermissionError:
                raise Exception(f"Permission denied: Cannot save file to {filepath}. Please check file permissions.")
            except Exception as e:
                raise Exception(f"Error saving Excel file: {str(e)}")
            
            logger.info(f"Excel bill generated: {filepath}")
            return str(filepath)
            
        except Exception as e:
            error_msg = f"Error generating Excel bill: {str(e)}"
            logger.error(error_msg)
            
            ai_suggestion = self.ai.handle_error(str(e), "Excel generation")
            logger.info(f"AI suggestion: {ai_suggestion}")
            
            raise Exception(f"{error_msg}\n\nAI Suggestion:\n{ai_suggestion}")
    
    def _create_summary_sheet(self, wb: openpyxl.Workbook, bill_data: Dict):
        """Create summary sheet with invoice details."""
        ws = wb.active
        ws.title = "Summary Bill"
        
        # Set column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        
        # Styles
        header_font = Font(name='Arial', size=14, bold=True)
        title_font = Font(name='Arial', size=16, bold=True)
        normal_font = Font(name='Arial', size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Company Header
        ws.merge_cells('A1:E1')
        ws['A1'] = "KRSNNAA DIAGNOSTICS LIMITED"
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells('A2:E2')
        ws['A2'] = "Medical Diagnostic Services"
        ws['A2'].font = header_font
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # Invoice Details
        invoice_number = self.invoice_generator.generate(bill_data['invoice_date'])
        
        ws['A4'] = "Invoice No:"
        ws['A4'].font = header_font
        ws['B4'] = invoice_number
        ws['B4'].font = normal_font
        
        ws['D4'] = "Date:"
        ws['D4'].font = header_font
        ws['E4'] = bill_data['invoice_date'].strftime('%d/%m/%Y')
        ws['E4'].font = normal_font
        
        ws['A5'] = "Bill To:"
        ws['A5'].font = header_font
        ws['B5'] = bill_data['center_name']
        ws['B5'].font = normal_font
        
        # Test Summary Table
        row = 8
        ws[f'A{row}'] = "S.No"
        ws[f'B{row}'] = "Test Type"
        ws[f'C{row}'] = "Count"
        ws[f'D{row}'] = "Rate per Test"
        ws[f'E{row}'] = "Net Amount"
        
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws[f'{col}{row}'].font = header_font
            ws[f'{col}{row}'].border = border
            ws[f'{col}{row}'].alignment = Alignment(horizontal='center')
        
        row += 1
        for i, test in enumerate(bill_data['test_summary'], 1):
            ws[f'A{row}'] = i
            ws[f'B{row}'] = test['test_type']
            ws[f'C{row}'] = test['count']
            ws[f'D{row}'] = f"₹{test['rate_per_test']:.2f}"
            ws[f'E{row}'] = f"₹{test['net_amount']:.2f}"
            
            for col in ['A', 'B', 'C', 'D', 'E']:
                ws[f'{col}{row}'].font = normal_font
                ws[f'{col}{row}'].border = border
                if col in ['C', 'D', 'E']:
                    ws[f'{col}{row}'].alignment = Alignment(horizontal='right')
            
            row += 1
        
        # Total
        ws[f'A{row}'] = ""
        ws[f'B{row}'] = "TOTAL"
        ws[f'C{row}'] = sum(test['count'] for test in bill_data['test_summary'])
        ws[f'D{row}'] = ""
        ws[f'E{row}'] = f"₹{bill_data['total_net']:.2f}"
        
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws[f'{col}{row}'].font = header_font
            ws[f'{col}{row}'].border = border
            if col in ['C', 'E']:
                ws[f'{col}{row}'].alignment = Alignment(horizontal='right')
        
        # Amount in Words
        row += 2
        ws[f'A{row}'] = "Amount in Words:"
        ws[f'A{row}'].font = header_font
        ws.merge_cells(f'B{row}:E{row}')
        ws[f'B{row}'] = self.amount_converter.convert(bill_data['total_net'])
        ws[f'B{row}'].font = normal_font
        
        # Narration
        row += 2
        ws[f'A{row}'] = "Narration:"
        ws[f'A{row}'].font = header_font
        
        period_text = ""
        if bill_data.get('period_start') and bill_data.get('period_end'):
            period_text = f" for the period {bill_data['period_start'].strftime('%d/%m/%Y')} to {bill_data['period_end'].strftime('%d/%m/%Y')}"
        
        ws.merge_cells(f'B{row}:E{row}')
        ws[f'B{row}'] = f"Medical diagnostic services provided{period_text}"
        ws[f'B{row}'].font = normal_font
        
        # Statutory Notes
        row += 3
        ws[f'A{row}'] = "Statutory Notes:"
        ws[f'A{row}'].font = header_font
        
        row += 1
        ws.merge_cells(f'A{row}:E{row}')
        ws[f'A{row}'] = "1. This invoice is exempt from GST as per Notification No. 12/2017-Central Tax (Rate) dated 28.06.2017"
        ws[f'A{row}'].font = normal_font
        
        row += 1
        ws.merge_cells(f'A{row}:E{row}')
        ws[f'A{row}'] = "2. TDS @ 2% u/s 194J of Income Tax Act, 1961 is applicable on professional/technical services"
        ws[f'A{row}'].font = normal_font
        
        # Bank Details
        row += 3
        ws[f'A{row}'] = "Bank Details:"
        ws[f'A{row}'].font = header_font
        
        row += 1
        ws[f'A{row}'] = "Bank Name:"
        ws[f'A{row}'].font = normal_font
        ws[f'B{row}'] = "HDFC Bank"
        ws[f'B{row}'].font = normal_font
        
        row += 1
        ws[f'A{row}'] = "Account No:"
        ws[f'A{row}'].font = normal_font
        ws[f'B{row}'] = "50200012345678"
        ws[f'B{row}'].font = normal_font
        
        row += 1
        ws[f'A{row}'] = "IFSC Code:"
        ws[f'A{row}'].font = normal_font
        ws[f'B{row}'] = "HDFC0001234"
        ws[f'B{row}'].font = normal_font
    
    def _create_detailed_sheet(self, wb: openpyxl.Workbook, bill_data: Dict):
        """Create detailed sheet with patient-wise information."""
        ws = wb.create_sheet("Detailed Bill")
        
        # Set column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 8
        ws.column_dimensions['F'].width = 30
        ws.column_dimensions['G'].width = 15
        
        # Styles
        header_font = Font(name='Arial', size=12, bold=True)
        normal_font = Font(name='Arial', size=10)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Header
        ws['A1'] = "SR.NO"
        ws['B1'] = "Date"
        ws['C1'] = "Patient Name"
        ws['D1'] = "Age"
        ws['E1'] = "Gender"
        ws['F1'] = "Test Name"
        ws['G1'] = "Value"
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            ws[f'{col}1'].font = header_font
            ws[f'{col}1'].border = border
            ws[f'{col}1'].alignment = Alignment(horizontal='center')
        
        # Patient details
        for i, patient in enumerate(bill_data['patient_details'], 1):
            row = i + 1
            ws[f'A{row}'] = patient['sr_no']
            ws[f'B{row}'] = patient['date']
            ws[f'C{row}'] = patient['patient_name']
            ws[f'D{row}'] = patient['age']
            ws[f'E{row}'] = patient['gender']
            ws[f'F{row}'] = patient['test_name']
            ws[f'G{row}'] = f"₹{patient['value']:.2f}"
            
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
                ws[f'{col}{row}'].font = normal_font
                ws[f'{col}{row}'].border = border
                if col == 'G':
                    ws[f'{col}{row}'].alignment = Alignment(horizontal='right')
    
    def generate_pdf_bill(self, bill_data: Dict) -> str:
        """Generate PDF bill with the same content as Excel."""
        try:
            filename = f"{bill_data['center_name'].replace(' ', '_').replace('/', '_').replace('\\', '_')}.pdf"
            filepath = self.output_dirs['pdf'] / filename
            
            # Ensure output directory exists
            self.output_dirs['pdf'].mkdir(parents=True, exist_ok=True)
            
            doc = SimpleDocTemplate(str(filepath), pagesize=A4)
            styles = getSampleStyleSheet()
            
            # Custom styles
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                spaceAfter=20,
                alignment=1  # Center
            )
            
            header_style = ParagraphStyle(
                'CustomHeader',
                parent=styles['Heading2'],
                fontSize=12,
                spaceAfter=10
            )
            
            normal_style = styles['Normal']
            
            # Build content
            story = []
            
            # Company Header
            story.append(Paragraph("KRSNNAA DIAGNOSTICS LIMITED", title_style))
            story.append(Paragraph("Medical Diagnostic Services", header_style))
            story.append(Spacer(1, 20))
            
            # Invoice Details
            invoice_number = self.invoice_generator.generate(bill_data['invoice_date'])
            
            invoice_data = [
                ['Invoice No:', invoice_number, 'Date:', bill_data['invoice_date'].strftime('%d/%m/%Y')],
                ['Bill To:', bill_data['center_name'], '', '']
            ]
            
            invoice_table = Table(invoice_data, colWidths=[1*inch, 2*inch, 1*inch, 1.5*inch])
            invoice_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            
            story.append(invoice_table)
            story.append(Spacer(1, 20))
            
            # Test Summary Table
            story.append(Paragraph("Test Summary", header_style))
            
            summary_data = [['S.No', 'Test Type', 'Count', 'Rate per Test', 'Net Amount']]
            for i, test in enumerate(bill_data['test_summary'], 1):
                summary_data.append([
                    str(i),
                    test['test_type'],
                    str(test['count']),
                    f"₹{test['rate_per_test']:.2f}",
                    f"₹{test['net_amount']:.2f}"
                ])
            
            # Add total row
            total_count = sum(test['count'] for test in bill_data['test_summary'])
            summary_data.append(['', 'TOTAL', str(total_count), '', f"₹{bill_data['total_net']:.2f}"])
            
            summary_table = Table(summary_data, colWidths=[0.5*inch, 2*inch, 0.8*inch, 1.2*inch, 1.2*inch])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('ALIGN', (2, 1), (-1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            
            story.append(summary_table)
            story.append(Spacer(1, 15))
            
            # Amount in Words
            story.append(Paragraph(f"Amount in Words: {self.amount_converter.convert(bill_data['total_net'])}", normal_style))
            story.append(Spacer(1, 15))
            
            # Narration
            period_text = ""
            if bill_data.get('period_start') and bill_data.get('period_end'):
                period_text = f" for the period {bill_data['period_start'].strftime('%d/%m/%Y')} to {bill_data['period_end'].strftime('%d/%m/%Y')}"
            
            story.append(Paragraph(f"Narration: Medical diagnostic services provided{period_text}", normal_style))
            story.append(Spacer(1, 20))
            
            # Statutory Notes
            story.append(Paragraph("Statutory Notes:", header_style))
            story.append(Paragraph("1. This invoice is exempt from GST as per Notification No. 12/2017-Central Tax (Rate) dated 28.06.2017", normal_style))
            story.append(Paragraph("2. TDS @ 2% u/s 194J of Income Tax Act, 1961 is applicable on professional/technical services", normal_style))
            story.append(Spacer(1, 20))
            
            # Bank Details
            story.append(Paragraph("Bank Details:", header_style))
            bank_data = [
                ['Bank Name:', 'HDFC Bank'],
                ['Account No:', '50200012345678'],
                ['IFSC Code:', 'HDFC0001234']
            ]
            
            bank_table = Table(bank_data, colWidths=[1.5*inch, 3*inch])
            bank_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            
            story.append(bank_table)
            
            # Build PDF
            try:
                doc.build(story)
            except PermissionError:
                raise Exception(f"Permission denied: Cannot save PDF file to {filepath}. Please check file permissions.")
            except Exception as e:
                raise Exception(f"Error generating PDF file: {str(e)}")
            
            logger.info(f"PDF bill generated: {filepath}")
            return str(filepath)
            
        except Exception as e:
            error_msg = f"Error generating PDF bill: {str(e)}"
            logger.error(error_msg)
            
            ai_suggestion = self.ai.handle_error(str(e), "PDF generation")
            logger.info(f"AI suggestion: {ai_suggestion}")
            
            raise Exception(f"{error_msg}\n\nAI Suggestion:\n{ai_suggestion}")
    
    def generate_all_bills(self, main_data_path: str, supporting_data_path: str = None) -> Dict[str, List[str]]:
        """Generate all bills (Excel and PDF) for all centers."""
        try:
            # Process billing data
            bills_data = self.process_billing_data(main_data_path, supporting_data_path)
            
            excel_files = []
            pdf_files = []
            
            # Generate bills for each center
            for bill_data in bills_data:
                logger.info(f"Generating bills for {bill_data['center_name']}")
                
                # Generate Excel bill
                excel_file = self.generate_excel_bill(bill_data)
                excel_files.append(excel_file)
                
                # Generate PDF bill
                pdf_file = self.generate_pdf_bill(bill_data)
                pdf_files.append(pdf_file)
            
            logger.info(f"Successfully generated {len(excel_files)} Excel bills and {len(pdf_files)} PDF bills")
            
            return {
                'excel_files': excel_files,
                'pdf_files': pdf_files,
                'total_centers': len(bills_data)
            }
            
        except Exception as e:
            error_msg = f"Error generating all bills: {str(e)}"
            logger.error(error_msg)
            
            ai_suggestion = self.ai.handle_error(str(e), "Bill generation process")
            logger.info(f"AI suggestion: {ai_suggestion}")
            
            raise Exception(f"{error_msg}\n\nAI Suggestion:\n{ai_suggestion}")

def interactive_ai_chat(ai_integration: AIIntegration):
    """Interactive AI chat mode for user assistance."""
    print("\n🤖 AI Assistant Mode")
    print("Ask me anything about the medical billing system!")
    print("Type 'exit' to return to the main program.\n")
    
    while True:
        try:
            user_input = input("You: ").strip()
            
            if user_input.lower() in ['exit', 'quit', 'q']:
                print("Exiting AI chat mode...\n")
                break
            
            if not user_input:
                continue
            
            print("AI: Thinking...")
            response = ai_integration.get_ai_response(user_input)
            print(f"AI: {response}\n")
            
        except KeyboardInterrupt:
            print("\nExiting AI chat mode...\n")
            break
        except Exception as e:
            print(f"Error: {e}\n")

def main():
    """Main function to run the medical billing application."""
    parser = argparse.ArgumentParser(
        description="Medical Test Billing Automation System",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python medical_billing_app.py --main-data BILL.xlsx
  python medical_billing_app.py --main-data BILL.xlsx --supporting-data sharing.xlsx
  python medical_billing_app.py --main-data BILL.xlsx --invoice-date 2024-01-15
  python medical_billing_app.py --ask
        """
    )
    
    parser.add_argument('--main-data', help='Path to main data Excel file (required unless using --ask)')
    parser.add_argument('--supporting-data', help='Path to supporting data Excel file (optional)')
    parser.add_argument('--invoice-date', help='Invoice date (YYYY-MM-DD format, default: today)')
    parser.add_argument('--period-start', help='Period start date (YYYY-MM-DD format)')
    parser.add_argument('--period-end', help='Period end date (YYYY-MM-DD format)')
    parser.add_argument('--invoice-sequence-start', type=int, default=1, help='Starting invoice sequence number')
    parser.add_argument('--ask', action='store_true', help='Enter AI chat mode for assistance')
    
    args = parser.parse_args()
    
    try:
        # Check if user wants AI chat mode
        if args.ask:
            ai = AIIntegration()
            interactive_ai_chat(ai)
            return
        
        # Validate main data file is provided
        if not args.main_data:
            print("❌ Error: --main-data argument is required!")
            print("Usage: python medical_billing_app.py --main-data BILL.xlsx")
            return
        
        # Validate input file exists
        if not os.path.exists(args.main_data):
            print(f"❌ Error: Main data file '{args.main_data}' not found!")
            return
        
        # Parse dates
        invoice_date = datetime.now()
        if args.invoice_date:
            try:
                invoice_date = datetime.strptime(args.invoice_date, '%Y-%m-%d')
            except ValueError:
                print("❌ Error: Invalid invoice date format. Use YYYY-MM-DD")
                return
        
        period_start = None
        if args.period_start:
            try:
                period_start = datetime.strptime(args.period_start, '%Y-%m-%d')
            except ValueError:
                print("❌ Error: Invalid period start date format. Use YYYY-MM-DD")
                return
        
        period_end = None
        if args.period_end:
            try:
                period_end = datetime.strptime(args.period_end, '%Y-%m-%d')
            except ValueError:
                print("❌ Error: Invalid period end date format. Use YYYY-MM-DD")
                return
        
        # Configuration
        config = {
            'invoice_date': invoice_date,
            'period_start': period_start,
            'period_end': period_end,
            'invoice_sequence_start': args.invoice_sequence_start
        }
        
        print("🏥 Medical Test Billing Automation System")
        print("=" * 50)
        print(f"📊 Main Data: {args.main_data}")
        if args.supporting_data:
            print(f"📋 Supporting Data: {args.supporting_data}")
        print(f"📅 Invoice Date: {invoice_date.strftime('%d/%m/%Y')}")
        if period_start and period_end:
            print(f"📅 Period: {period_start.strftime('%d/%m/%Y')} to {period_end.strftime('%d/%m/%Y')}")
        print()
        
        # Initialize processor
        processor = MedicalBillingProcessor(config)
        
        # Generate all bills
        print("🔄 Processing billing data...")
        result = processor.generate_all_bills(args.main_data, args.supporting_data)
        
        print("\n✅ Billing completed successfully!")
        print(f"📁 Generated {result['total_centers']} center bills")
        print(f"📊 Excel files: {len(result['excel_files'])}")
        print(f"📄 PDF files: {len(result['pdf_files'])}")
        print("\n📂 Output locations:")
        print(f"   Excel: bills/excel/")
        print(f"   PDF: bills/pdf/")
        
    except Exception as e:
        print(f"\n❌ Error: {str(e)}")
        
        # Offer AI assistance
        try:
            ai = AIIntegration()
            print("\n🤖 Would you like AI assistance? (y/n): ", end='')
            response = input().strip().lower()
            
            if response in ['y', 'yes']:
                print("\n" + "="*50)
                ai_suggestion = ai.handle_error(str(e), "Main application execution")
                print(f"AI Suggestion:\n{ai_suggestion}")
        except:
            pass

if __name__ == "__main__":
    main() 