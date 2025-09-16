# --- Imports and App Initialization ---
import os
import logging
from flask import Flask, render_template, request, redirect, url_for, flash, session, send_file, jsonify
from functools import wraps
# ...existing code...

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)
app.secret_key = 'your-secret-key-here'
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# --- Login Required Decorator ---
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Login required', 'error')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# --- Advanced Dashboard & Analytics Route ---
@app.route('/dashboard')
@login_required
def dashboard():
    try:
        # Filters
        month = request.args.get('month')
        category_name = request.args.get('category')
        query = Bill.query
        if month:
            query = query.filter(Bill.month == month)
        if category_name:
            category = BillCategory.query.filter_by(name=category_name).first()
            if category:
                query = query.filter(Bill.category_id == category.id)
        bills_db = query.all()
        bills = [b.bill_data for b in bills_db]

        # --- Metrics ---
        total_revenue = sum(b.get('total_rate', 0) for b in bills)
        total_outstanding = sum(
            b.get('total_rate', 0) - sum(b.get('payment_details', {}).values())
            if b.get('payment_details') else b.get('total_rate', 0)
            for b in bills
        )
        num_bills = len(bills)
        # Top center this month
        top_center = None
        if bills:
            center_totals = {}
            for b in bills:
                center = b.get('centre_name')
                center_totals[center] = center_totals.get(center, 0) + b.get('total_rate', 0)
            top_center = max(center_totals, key=center_totals.get)

        # --- Monthly Billing Trend ---
        # Group by month (YYYY-MM)
        from collections import defaultdict
        monthly_trend = defaultdict(float)
        for b in Bill.query.all():
            m = b.month
            monthly_trend[m] += b.bill_data.get('total_rate', 0)
        monthly_trend = sorted(monthly_trend.items())

        # --- Top 5 Centers ---
        center_totals = defaultdict(float)
        for b in bills:
            center = b.get('centre_name')
            center_totals[center] += b.get('total_rate', 0)
        top_centers = sorted(center_totals.items(), key=lambda x: x[1], reverse=True)[:5]

        # --- Category Distribution ---
        category_totals = defaultdict(float)
        for b in bills_db:
            cat = b.category.name if b.category else 'Unknown'
            category_totals[cat] += b.bill_data.get('total_rate', 0)
        category_dist = sorted(category_totals.items())

        # --- Outstanding vs Paid ---
        paid = sum(
            sum(b.get('payment_details', {}).values())
            if b.get('payment_details') else 0
            for b in bills
        )
        outstanding = total_revenue - paid

        return render_template(
            'dashboard.html',
            total_revenue=total_revenue,
            total_outstanding=total_outstanding,
            num_bills=num_bills,
            top_center=top_center,
            monthly_trend=monthly_trend,
            top_centers=top_centers,
            category_dist=category_dist,
            paid=paid,
            outstanding=outstanding,
            app=app
        )
    except Exception as e:
        logger.error(f"Error in dashboard route: {e}")
        flash('An error occurred while loading dashboard', 'error')
        return redirect(url_for('index'))
# --- Audit Logs Route (admin only) ---
# (Route is defined after permission decorators below)
from flask import Flask, render_template, request, jsonify, send_file, redirect, url_for, flash, session
import pandas as pd
import os
from datetime import datetime
import json
from werkzeug.utils import secure_filename
import tempfile
import zipfile
from io import BytesIO
import logging
import requests
from dotenv import load_dotenv
import re
from utils import AmountToWords, InvoiceNumberGenerator, AIIntegration, safe_float_conversion, safe_int_conversion, safe_date_conversion
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
# SQLAlchemy/PostgreSQL imports
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import func
from openpyxl.cell import MergedCell
from utils_auth import hash_password, verify_password
from utils_email import send_email_with_attachment

# Load environment variables
load_dotenv()

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)
app.secret_key = 'your-secret-key-here'
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# SQLAlchemy/PostgreSQL config
app.config['SQLALCHEMY_DATABASE_URI'] = os.getenv('DATABASE_URL', 'postgresql://postgres:postgres@localhost:5432/billing_db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)

# Ensure upload folder exists
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

ALLOWED_EXTENSIONS = {'xlsx', 'xls', 'xlsm'}

# Initialize global objects
amount_converter = AmountToWords()
invoice_generator = InvoiceNumberGenerator()
ai_integration = AIIntegration()

# --- Update Bill Payment Info ---
@app.route('/bill/<int:bill_index>/update_payment', methods=['POST'])
def update_bill_payment(bill_index):
    try:
        if not hasattr(app, 'bills') or not app.bills:
            flash('No bills available', 'error')
            return redirect(url_for('bills'))
        if bill_index < 0 or bill_index >= len(app.bills):
            flash('Bill not found', 'error')
            return redirect(url_for('bills'))

        # Get payment info from form
        payment_mode = request.form.get('payment_mode')
        payment_methods = request.form.getlist('payment_method[]')
        payment_amounts = request.form.getlist('payment_amount[]')
        payment_details = {}
        for method, amount in zip(payment_methods, payment_amounts):
            if method and amount:
                try:
                    payment_details[method] = float(amount)
                except Exception:
                    continue

        # Update in DB (assume Bill model is in DB, not just app.bills)
        bill_obj = Bill.query.filter_by(bill_number=app.bills[bill_index]['bill_number']).first()
        if bill_obj:
            bill_obj.payment_mode = payment_mode
            bill_obj.payment_details = payment_details
            db.session.commit()
            # Update in-memory bill if needed
            app.bills[bill_index]['payment_mode'] = payment_mode
            app.bills[bill_index]['payment_details'] = payment_details
            flash('Payment info updated.', 'success')
        else:
            flash('Bill not found in database.', 'error')
        return redirect(url_for('view_bill', bill_index=bill_index))
    except Exception as e:
        logger.error(f"Error in update_bill_payment: {e}")
        flash('An error occurred while updating payment info', 'error')
        return redirect(url_for('view_bill', bill_index=bill_index))
# Load environment variables
from utils_auth import hash_password, verify_password
from flask import session
# --- User Registration Route ---
@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form.get('username')
        email = request.form.get('email')
        password = request.form.get('password')
        role = request.form.get('role', 'staff')
        if not username or not email or not password:
            flash('All fields are required', 'error')
            return render_template('register.html')
        if User.query.filter((User.username == username) | (User.email == email)).first():
            flash('Username or email already exists', 'error')
            return render_template('register.html')
        user = User(
            username=username,
            email=email,
            password_hash=hash_password(password),
            role=role
        )
        db.session.add(user)
        db.session.commit()
        flash('Registration successful. Please log in.', 'success')
        return redirect(url_for('login'))
    return render_template('register.html')

# --- User Login Route ---
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        identifier = request.form.get('identifier')  # username or email
        password = request.form.get('password')
        user = User.query.filter((User.username == identifier) | (User.email == identifier)).first()
        if user and verify_password(user.password_hash, password):
            session['user_id'] = user.id
            session['username'] = user.username
            session['role'] = user.role
            flash('Login successful', 'success')
            return redirect(url_for('index'))
        else:
            flash('Invalid credentials', 'error')
    return render_template('login.html')

# --- User Logout Route ---
@app.route('/logout')
def logout():
    session.clear()
    flash('Logged out successfully', 'success')
    return redirect(url_for('login'))
# Load environment variables
from utils_email import send_email_with_attachment
# --- Email Bill Endpoint ---
@app.route('/email_bill/<int:bill_index>', methods=['POST'])
def email_bill(bill_index):
    try:
        if not hasattr(app, 'bills') or not app.bills:
            flash('No bills available', 'error')
            return redirect(url_for('bills'))
        if bill_index < 0 or bill_index >= len(app.bills):
            flash('Bill not found', 'error')
            return redirect(url_for('bills'))

        bill = app.bills[bill_index]
        fmt = request.form.get('format', 'pdf').lower()
        email_to = request.form.get('email_to')
        smtp_provider = request.form.get('smtp_provider', 'gmail')
        if not email_to:
            flash('Recipient email required', 'error')
            return redirect(url_for('view_bill', bill_index=bill_index))

        # Prepare attachment
        safe_center_name = str(bill['centre_name']).replace(" ", "_").replace("/", "_")
        subject = f"Bill {bill['bill_number']} - {bill['centre_name']}"
        body = f"Please find attached the bill {bill['bill_number']} for {bill['centre_name']}."
        attachment_bytes = None
        attachment_filename = None
        if fmt == 'excel':
            if bill.get("center_type") == "HLM":
                excel_path = generate_hlm_excel_from_template(bill, bill.get('center_rows_data', []))
                with open(excel_path, 'rb') as f:
                    attachment_bytes = f.read()
                attachment_filename = f"{safe_center_name}.xlsm"
            else:
                import pandas as pd
                from io import BytesIO
                buffer = BytesIO()
                df = pd.DataFrame(bill['test_items'])
                df.to_excel(buffer, index=False)
                buffer.seek(0)
                attachment_bytes = buffer.read()
                attachment_filename = f"{safe_center_name}.xlsx"
        elif fmt == 'pdf':
            html_content = render_template('bill_pdf.html', bill=bill)
            from io import BytesIO
            pdf_buffer = BytesIO()
            try:
                import pdfkit
                pdf = pdfkit.from_string(html_content, False)
                pdf_buffer.write(pdf)
            except Exception:
                from xhtml2pdf import pisa
                pisa_status = pisa.CreatePDF(html_content, dest=pdf_buffer)
                if pisa_status.err:
                    flash('PDF generation failed', 'error')
                    return redirect(url_for('view_bill', bill_index=bill_index))
            pdf_buffer.seek(0)
            attachment_bytes = pdf_buffer.read()
            attachment_filename = f"{safe_center_name}.pdf"
        else:
            html_content = render_template('bill_pdf.html', bill=bill)
            attachment_bytes = html_content.encode('utf-8')
            attachment_filename = f"{safe_center_name}.html"

        send_email_with_attachment(
            subject=subject,
            body=body,
            to_emails=[email_to],
            attachment_bytes=attachment_bytes,
            attachment_filename=attachment_filename,
            smtp_provider=smtp_provider
        )
        flash(f'Bill emailed to {email_to} via {smtp_provider}', 'success')
        return redirect(url_for('view_bill', bill_index=bill_index))
    except Exception as e:
        logger.error(f"Error in email_bill: {e}")
        flash('An error occurred while emailing the bill', 'error')
        return redirect(url_for('view_bill', bill_index=bill_index))
# Load environment variables
from flask import Flask, render_template, request, jsonify, send_file, redirect, url_for, flash
import pandas as pd
import os
from datetime import datetime
import json
from werkzeug.utils import secure_filename
import tempfile
import zipfile
from io import BytesIO
import logging
import requests
from dotenv import load_dotenv
import re
from utils import AmountToWords, InvoiceNumberGenerator, AIIntegration, safe_float_conversion, safe_int_conversion, safe_date_conversion
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
# SQLAlchemy/PostgreSQL imports
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import func
from openpyxl.cell import MergedCell

# Load environment variables
load_dotenv()

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)
app.secret_key = 'your-secret-key-here'
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# SQLAlchemy/PostgreSQL config
app.config['SQLALCHEMY_DATABASE_URI'] = os.getenv('DATABASE_URL', 'postgresql://postgres:postgres@localhost:5432/billing_db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)

# Ensure upload folder exists
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

ALLOWED_EXTENSIONS = {'xlsx', 'xls', 'xlsm'}

# Initialize global objects
amount_converter = AmountToWords()
invoice_generator = InvoiceNumberGenerator()
ai_integration = AIIntegration()

# --- Database Models ---
class BillCategory(db.Model):
    __tablename__ = 'bill_categories'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(64), unique=True, nullable=False)
    description = db.Column(db.String(256))
    bills = db.relationship('Bill', backref='category', lazy=True)

class UploadedFile(db.Model):
    __tablename__ = 'uploaded_files'
    id = db.Column(db.Integer, primary_key=True)
    filename = db.Column(db.String(256), nullable=False)
    upload_time = db.Column(db.DateTime, default=datetime.utcnow)
    file_path = db.Column(db.String(512), nullable=False)
    bills = db.relationship('Bill', backref='uploaded_file', lazy=True)

class Bill(db.Model):
    __tablename__ = 'bills'
    id = db.Column(db.Integer, primary_key=True)
    bill_number = db.Column(db.String(64), unique=True, nullable=False)
    center_name = db.Column(db.String(128), nullable=False)
    month = db.Column(db.String(16), nullable=False)  # e.g. '2025-08'
    category_id = db.Column(db.Integer, db.ForeignKey('bill_categories.id'), nullable=False)
    uploaded_file_id = db.Column(db.Integer, db.ForeignKey('uploaded_files.id'), nullable=True)
    bill_data = db.Column(db.JSON, nullable=False)  # Store bill details as JSON
    status = db.Column(db.String(20), nullable=False, default='pending')  # pending, paid, rejected, in_review, cancelled
    payment_mode = db.Column(db.String(50), nullable=True)  # cash, phonepay, mobivik, hdfc_ac, yesbank_ac, manglam, card, etc.
    payment_details = db.Column(db.JSON, nullable=True)  # breakdown: {"manglam": 1000, "hdfc_ac": 500, ...}
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    def __repr__(self):
        return f'<Bill {self.bill_number} - {self.center_name}>'

class User(db.Model):
    __tablename__ = 'users'
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(64), unique=True, nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(128), nullable=False)
    role = db.Column(db.String(16), nullable=False, default='staff')  # 'admin' or 'staff'
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    can_edit_bills = db.Column(db.Boolean, default=False)
    can_delete_bills = db.Column(db.Boolean, default=False)

    def __repr__(self):
        return f'<User {self.username} ({self.role})>'

from flask import Flask, render_template, request, jsonify, send_file, redirect, url_for, flash
import pandas as pd
import os
from datetime import datetime
import json
from werkzeug.utils import secure_filename
import tempfile
import zipfile
from io import BytesIO
import logging
import requests
from dotenv import load_dotenv
import re
from utils import AmountToWords, InvoiceNumberGenerator, AIIntegration, safe_float_conversion, safe_int_conversion, safe_date_conversion
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
# SQLAlchemy/PostgreSQL imports
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import func
from openpyxl.cell import MergedCell

# Load environment variables
load_dotenv()

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)
app.secret_key = 'your-secret-key-here'
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# SQLAlchemy/PostgreSQL config
app.config['SQLALCHEMY_DATABASE_URI'] = os.getenv('DATABASE_URL', 'postgresql://postgres:postgres@localhost:5432/billing_db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)

# Ensure upload folder exists
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

ALLOWED_EXTENSIONS = {'xlsx', 'xls', 'xlsm'}

...

# Load environment variables
load_dotenv()

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)
app.secret_key = 'your-secret-key-here'
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# SQLAlchemy/PostgreSQL config
app.config['SQLALCHEMY_DATABASE_URI'] = os.getenv('DATABASE_URL', 'postgresql://postgres:postgres@localhost:5432/billing_db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)

# --- Database Models ---
class BillCategory(db.Model):
    __tablename__ = 'bill_categories'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(64), unique=True, nullable=False)
    description = db.Column(db.String(256))
    bills = db.relationship('Bill', backref='category', lazy=True)

class UploadedFile(db.Model):
    __tablename__ = 'uploaded_files'
    id = db.Column(db.Integer, primary_key=True)
    filename = db.Column(db.String(256), nullable=False)
    upload_time = db.Column(db.DateTime, default=datetime.utcnow)
    file_path = db.Column(db.String(512), nullable=False)
    bills = db.relationship('Bill', backref='uploaded_file', lazy=True)


# --- Database Models ---
class BillCategory(db.Model):
    __tablename__ = 'bill_categories'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(64), unique=True, nullable=False)
    description = db.Column(db.String(256))
    bills = db.relationship('Bill', backref='category', lazy=True)

class UploadedFile(db.Model):
    __tablename__ = 'uploaded_files'
    id = db.Column(db.Integer, primary_key=True)
    filename = db.Column(db.String(256), nullable=False)
    upload_time = db.Column(db.DateTime, default=datetime.utcnow)
    file_path = db.Column(db.String(512), nullable=False)
    bills = db.relationship('Bill', backref='uploaded_file', lazy=True)

class Bill(db.Model):
    __tablename__ = 'bills'
    id = db.Column(db.Integer, primary_key=True)
    bill_number = db.Column(db.String(64), unique=True, nullable=False)
    center_name = db.Column(db.String(128), nullable=False)
    month = db.Column(db.String(16), nullable=False)  # e.g. '2025-08'
    category_id = db.Column(db.Integer, db.ForeignKey('bill_categories.id'), nullable=False)
    uploaded_file_id = db.Column(db.Integer, db.ForeignKey('uploaded_files.id'), nullable=True)
    bill_data = db.Column(db.JSON, nullable=False)  # Store bill details as JSON
    status = db.Column(db.String(20), nullable=False, default='pending')  # pending, paid, rejected, in_review, cancelled
    payment_mode = db.Column(db.String(50), nullable=True)  # cash, phonepay, mobivik, hdfc_ac, yesbank_ac, manglam, card, etc.
    payment_details = db.Column(db.JSON, nullable=True)  # breakdown: {"manglam": 1000, "hdfc_ac": 500, ...}
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    def __repr__(self):
        return f'<Bill {self.bill_number} - {self.center_name}>'

class User(db.Model):
    __tablename__ = 'users'
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(64), unique=True, nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(128), nullable=False)
    role = db.Column(db.String(16), nullable=False, default='staff')  # 'admin' or 'staff'
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    can_edit_bills = db.Column(db.Boolean, default=False)
    can_delete_bills = db.Column(db.Boolean, default=False)

    def __repr__(self):
        return f'<User {self.username} ({self.role})>'
from flask import Flask, render_template, request, jsonify, send_file, redirect, url_for, flash
import pandas as pd
import os
from datetime import datetime
import json
from werkzeug.utils import secure_filename
import tempfile
import zipfile
from io import BytesIO
import logging
import requests
from dotenv import load_dotenv
import re
from utils import AmountToWords, InvoiceNumberGenerator, AIIntegration, safe_float_conversion, safe_int_conversion, safe_date_conversion
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
# SQLAlchemy/PostgreSQL imports
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import func
from openpyxl.cell import MergedCell

# Load environment variables
load_dotenv()

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


app = Flask(__name__)
app.secret_key = 'your-secret-key-here'
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# SQLAlchemy/PostgreSQL config
app.config['SQLALCHEMY_DATABASE_URI'] = os.getenv('DATABASE_URL', 'postgresql://postgres:postgres@localhost:5432/billing_db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)

# Ensure upload folder exists
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

ALLOWED_EXTENSIONS = {'xlsx', 'xls', 'xlsm'}


# Initialize global objects
amount_converter = AmountToWords()
invoice_generator = InvoiceNumberGenerator()
ai_integration = AIIntegration()

# 🔹 Helper: Generate HLM Excel from Template
def generate_hlm_excel_from_template(bill, center_rows, template_path="HLM_Template.xlsm", output_folder="hlm_bills"):
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    # ✅ VBA/macros को preserve करने के लिए keep_vba=True
    wb = load_workbook(template_path, keep_vba=True)

    # ================= Sheet1 (Invoice) =================
    ws1 = wb["Invoice"]
    ws1['B12'] = bill['centre_name']
    today = datetime.today()
    third_of_current_month = today.replace(day=3)
    ws1['G11'] = third_of_current_month.strftime("%Y-%m-%d")  # Or any format you like
    ws1['G12'] = bill['bill_number']
    ws1['G13'] = "August-2025"

    # -------- Modality count logic (only this center) --------
    pathology_count = sum(
        1 for row in center_rows if str(row.get("Modality", "")).strip().lower() == "pathology"
    )
    radiology_count = sum(
        1 for row in center_rows if str(row.get("Modality", "")).strip().lower() in ["radiology", "nuclear"]
    )

    if pathology_count > 0:
        ws1["B19"] = "Pathology Investigation"
        ws1["E19"] = pathology_count

    if radiology_count > 0:
        ws1["B20"] = "Radiology Investigation"
        ws1["E20"] = radiology_count

    # Totals
    total_mrp = sum(t['mrp'] for t in bill['test_items'])
    total_sharing = sum(t['sharing_amount'] for t in bill['test_items'])
    ws1['G19'] = total_mrp
    ws1['G27'] = total_sharing

    # ================= Sheet2 (Detailed) =================
   

    ws2 = wb["Detailed"]

    # 🔹 Existing merged cells को unmerge करें
     # openpyxl के MergedCellRange objects का उपयोग करें
    merged_cells_ranges = list(ws2.merged_cells.ranges) 
    for merged_range in merged_cells_ranges:
        ws2.unmerge_cells(str(merged_range))

# 🔹 पुराना data clear (पूरी sheet)
    for row in ws2.iter_rows():
        for cell in row:
            cell.value = None
# Styles
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    total_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    bold_font = Font(bold=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
# Row 1 → Center name (merge A1:H1)
    ws2.merge_cells("A1:H1")
    ws2["A1"] = f"{bill['centre_name']}"
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2["A1"].font = Font(bold=True, size=14)

# Row 2 → Headings
    headings = [
        "RegisteredDate",
        "PatientName",
        "VisitCode",
        "TestName",
        "TestType",
        "MRP",
        "SharingAmount",
        "Net Amount"
        ]
    for col, heading in enumerate(headings, 1):
        cell = ws2.cell(row=2, column=col, value=heading)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

# Row 3 से details भरना
    for i, item in enumerate(bill['test_items'], start=3):
        values = [
            item.get('registered_date', ''),
            item.get('patient_name', ''),
            item.get('visit_code', ''),
            item.get('test_name', ''),
            item.get('modality', ''),
            item.get('mrp', 0),
            item.get('sharing_amount', 0),
            item.get('rate', 0)
        ]
        for col, val in enumerate(values, 1):
            cell = ws2.cell(row=i, column=col, value=val)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        # Totals row
    last_row = len(bill['test_items']) + 3
    ws2[f"E{last_row}"] = "TOTAL"
    ws2[f"E{last_row}"].font = bold_font
    ws2[f"E{last_row}"].fill = total_fill
    ws2[f"E{last_row}"].alignment = Alignment(horizontal="center")

    ws2[f"F{last_row}"] = total_mrp
    ws2[f"G{last_row}"] = total_sharing
    ws2[f"H{last_row}"] = bill['total_rate']

    for col in ["F", "G", "H"]:
        cell = ws2[f"{col}{last_row}"]
        cell.font = bold_font
        cell.fill = total_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # 🔹 Auto-adjust column widths for Detailed sheet (ws2) safely
    column_max_lengths = {}
    for row_idx in range(1, ws2.max_row + 1):
        for col_idx in range(1, ws2.max_column + 1):
            cell = ws2.cell(row=row_idx, column=col_idx)
            if cell.value is not None and not isinstance(cell, MergedCell): # MergedCell को छोड़ दें
                current_length = len(str(cell.value))
                col_letter = get_column_letter(col_idx)
                column_max_lengths[col_letter] = max(column_max_lengths.get(col_letter, 0), current_length)
  
    for col_letter, max_length in column_max_lengths.items():
        ws2.column_dimensions[col_letter].width = max_length + 2

    # Save file
    safe_center_name = str(bill['centre_name']).replace(" ", "_").replace("/", "_")
    output_path = os.path.join(output_folder, f"{safe_center_name}.xlsm")
    wb.save(output_path)

    return output_path

    # ---------------- Existing Code Below ---------------- #

def allowed_file(filename):
    """Check if file extension is allowed"""
    if not filename or '.' not in filename:
        return False
    return filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def validate_excel_data(df):
    """Validate Excel data structure and required columns"""
    required_columns = ['RegisteredDate', 'PatientVisitCode', 'PatientName', 'TEST NAME', 'MRP', 'CentreTestRate', 'CENTER NAME']
    
    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        return False, f"Missing required columns: {', '.join(missing_columns)}"
    
    # Check for empty dataframe
    if df.empty:
        return False, "Excel file is empty"
    
    # Check for required data in key columns
    if df['CENTER NAME'].isna().all():
        return False, "No center names found in the data"
    
    if df['MRP'].isna().all():
        return False, "No MRP values found in the data"
    
    return True, None



def process_excel_file(file_path):
    """Process Excel file and return billing data grouped by Center Name, with MobileNumber filtering."""

# --- Audit Log Model ---
class AuditLog(db.Model):
    __tablename__ = 'audit_logs'
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=True)
    action = db.Column(db.String(64), nullable=False)
    bill_id = db.Column(db.Integer, db.ForeignKey('bills.id'), nullable=True)
    timestamp = db.Column(db.DateTime, default=datetime.utcnow)
    details = db.Column(db.Text, nullable=True)

    user = db.relationship('User', backref='audit_logs')
    bill = db.relationship('Bill', backref='audit_logs')

    def __repr__(self):
        return f'<AuditLog {self.action} by {self.user_id} on {self.bill_id}>'
from flask import abort
# --- Helper: Permission Check Decorators ---

# --- Audit Logs Route (admin only) ---
# Place after permission check decorators so decorators are defined

from functools import wraps

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Login required', 'error')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if session.get('role') != 'admin':
            flash('Admin access required', 'error')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

def can_edit_bills_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        user = User.query.get(session.get('user_id'))
        if not user or not (user.role == 'admin' or user.can_edit_bills):
            flash('You do not have permission to edit bills', 'error')
            return redirect(url_for('bills'))
        return f(*args, **kwargs)
    return decorated_function

def can_delete_bills_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        user = User.query.get(session.get('user_id'))
        if not user or not (user.role == 'admin' or user.can_delete_bills):
            flash('You do not have permission to delete bills', 'error')
            return redirect(url_for('bills'))
        return f(*args, **kwargs)
    return decorated_function
# --- Edit Bill Endpoint ---
@app.route('/bill/<int:bill_id>/edit', methods=['GET', 'POST'])
@login_required
@can_edit_bills_required
def edit_bill(bill_id):
    bill = Bill.query.get_or_404(bill_id)
    if request.method == 'POST':
        # Example: allow editing status and payment_mode only
        status = request.form.get('status')
        payment_mode = request.form.get('payment_mode')
        if status:
            bill.status = status
        if payment_mode:
            bill.payment_mode = payment_mode
        db.session.commit()
        # Audit log
        log = AuditLog(user_id=session.get('user_id'), action='edit', bill_id=bill.id, details=f"Status: {status}, Payment mode: {payment_mode}")
        db.session.add(log)
        db.session.commit()
        flash('Bill updated successfully', 'success')
        return redirect(url_for('view_bill', bill_index=0))  # TODO: update to correct index logic
    return render_template('edit_bill.html', bill=bill, app=app)

# --- Delete Bill Endpoint ---
@app.route('/bill/<int:bill_id>/delete', methods=['POST'])
@login_required
@can_delete_bills_required
def delete_bill(bill_id):
    bill = Bill.query.get_or_404(bill_id)
    db.session.delete(bill)
    db.session.commit()
    # Audit log
    log = AuditLog(user_id=session.get('user_id'), action='delete', bill_id=bill.id, details='Bill deleted')
    db.session.add(log)
    db.session.commit()
    flash('Bill deleted successfully', 'success')
    return redirect(url_for('bills'))

    # ...existing code...
def get_hlm_centers():
    """Get list of HLM centers"""
    hlm_centers = [
        'Manglam_Diagnostics_Agroha',
        'Hansi_Lab_MANGLAM',
        'JIND_MANGLAM_LAB_HISAR',
        'Narwana_Manglam_Lab',
        'Sanyam_Lab,_Inside_Satija_Healthcare,_H..',
        'Shri_Bala_JI_Lab_Isherwal,_Miran',
        'Vikash_Laboratory,_Java_Hospital,_Tosha..'
    ]
    return hlm_centers

def get_b2b_centers():
    """Get list of B2B centers (non-HLM)"""
    all_centers = [
        'Aarogya_Hospital_Hisar',
        'AMANDEEP_HOSPITAL',
        'AMARAVATI_HOSPITAL',
        'APEX_DIAGNOSTICS',
        'Barwala_Jansevarth_Lab,_Barwala',
        'CITY_CENTER_HISAR',
        'DR_ANKIT_GOYAL',
        'DR._RAJESH_MEHTA',
        'ECHS',
        'Elora_Dass_Gupta',
        'Fatehabad_Manglam_Diagnostices',
        'Geetanjali_Hospital',
        'GOBIND_NARSING_HOME',
        'GUPTA_NEWTON_HOSPITAL',
        'Guru_Jambheshwar_Multispeciality_Hosp....',
        'HISAR_DIAGNOSTICS_JHAJHPUL',
        'HISAR_GESTRO_HOSPITAL',
        'Hisar_Hospital_Nursery_Inside_Hsr_Hosp..',
        'HOLY_HELP_HOSPITAL',
        'INSURANCE,_HISAR',
        'JANKI_HOSPITAL',
        'LIFE_LINE_HOSPITAL',
        'MEYANSH_HOSPITAL',
        'Navjeevan_Hospital',
        'Onquest_Laboratories_Ltd..',
        'Pathkind_Diagnostics',
        'Ram_Niwas\'s_Centre',
        'Ravindra_Hospital',
        'RMCT_TOHANA',
        'SACHIN_MITTAL',
        'SADBHAVNA_HOSPITAL',
        'Sai_Hospital',
        'Sapra_Hospital,_Hisar',
        'SARVODYA_HOSPITAL',
        'SHANI_MANAV_SEVA_TRUST',
        'SHANTI_GI_HOSPITAL',
        'SHARDHA_HOSPITAL',
        'Shree_Krishna_Pranami_Multi_speciality_H..'
    ]
    hlm_centers = get_hlm_centers()
    b2b_centers = [center for center in all_centers if center not in hlm_centers]
    return b2b_centers

@app.route('/')
def index():
    try:
        return render_template('index.html', app=app)
    except Exception as e:
        logger.error(f"Error in index route: {e}")
        flash('An error occurred while loading the page', 'error')
    return render_template('index.html', app=app)

@app.route('/upload', methods=['POST'])
def upload_file():
    try:
        if 'file' not in request.files:
            flash('No file selected', 'error')
            return redirect(url_for('index'))
        file = request.files['file']
        if file.filename == '':
            flash('No file selected', 'error')
            return redirect(url_for('index'))
        if not file or not allowed_file(file.filename):
            flash('Invalid file type. Please upload an Excel file (.xlsx or .xls)', 'error')
            return redirect(url_for('index'))
        try:
            filename = secure_filename(file.filename)
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(file_path)
            # Upload to Google Drive (optional: set your folder ID)
            try:
                from utils_gdrive import upload_file_to_gdrive
                gdrive_file_id = upload_file_to_gdrive(file_path, drive_folder_id=None)  # Set folder ID if needed
                logger.info(f"Uploaded to Google Drive, file ID: {gdrive_file_id}")
            except Exception as gdrive_exc:
                logger.error(f"Google Drive upload failed: {gdrive_exc}")
        except Exception as e:
            logger.error(f"Error saving file: {e}")
            flash('Error saving uploaded file', 'error')
            return redirect(url_for('index'))
        df, error = process_excel_file(file_path)
        if error:
            ai_suggestion = ai_integration.handle_error(error, "File upload and processing")
            flash(f'Error processing file: {error}\n\nAI Suggestion:\n{ai_suggestion}', 'error')
            return redirect(url_for('index'))
        if df is None or df.empty:
            flash('No bills could be generated from the uploaded file', 'error')
            return redirect(url_for('index'))
        # Store uploaded file info in DB
        uploaded_file = UploadedFile(filename=filename, file_path=file_path)
        db.session.add(uploaded_file)
        db.session.commit()
        app.df = df
        # --- Audit log for upload ---
        user_id = session.get('user_id') if 'user_id' in session else None
        log = AuditLog(user_id=user_id, action='upload', bill_id=None, details=f"Uploaded file: {filename}")
        db.session.add(log)
        db.session.commit()
        flash(f'Successfully uploaded {filename}', 'success')
        return redirect(url_for('index'))
    except Exception as e:
        logger.error(f"Error in upload_file: {e}")
        ai_suggestion = ai_integration.handle_error(str(e), "File upload process")
        flash(f'An unexpected error occurred while processing the file\n\nAI Suggestion:\n{ai_suggestion}', 'error')
        return redirect(url_for('index'))

@app.route('/generate_all_bills')
def generate_all_bills():
    """Generate bills for all centers"""
    try:
        if not hasattr(app, 'bills') or not app.bills:
            flash('No bills available. Please upload an Excel file first.', 'error')
            return redirect(url_for('index'))
        # --- Audit log for bill generation ---
        user_id = session.get('user_id') if 'user_id' in session else None
        log = AuditLog(user_id=user_id, action='generate_all_bills', bill_id=None, details='Generated all bills')
        db.session.add(log)
        db.session.commit()
        flash('All bills generated successfully!', 'success')
        return redirect(url_for('bills'))
    except Exception as e:
        logger.error(f"Error in generate_all_bills: {e}")
        flash('An error occurred while generating bills', 'error')
        return redirect(url_for('bills'))

@app.route('/generate_manual_bill', methods=['GET', 'POST'])
def generate_manual_bill():
    """Show manual bill generation page or process single bill generation"""
    try:
        if not hasattr(app, 'bills') or not app.bills:
            flash('No bills available. Please upload an Excel file first.', 'error')
            return redirect(url_for('index'))
        
        if request.method == 'POST':
            center_name = request.form.get('center_name')
            if not center_name:
                flash('Please select a center', 'error')
                return redirect(url_for('generate_manual_bill'))
            
            # Find the selected bill
            selected_bill = None
            for bill in app.bills:
                if bill['centre_name'] == center_name:
                    selected_bill = bill
                    break
            
            if selected_bill:
                # Create a new bills list with only the selected bill
                app.bills = [selected_bill]
                # --- Audit log for manual bill generation ---
                user_id = session.get('user_id') if 'user_id' in session else None
                log = AuditLog(user_id=user_id, action='generate_manual_bill', bill_id=None, details=f'Generated manual bill for {center_name}')
                db.session.add(log)
                db.session.commit()
                flash(f'Generated bill for {center_name}', 'success')
                return redirect(url_for('bills'))
            else:
                flash('Selected center not found', 'error')
                return redirect(url_for('generate_manual_bill'))
        
        return render_template('manual_bill.html', bills=app.bills, app=app)
    except Exception as e:
        logger.error(f"Error in generate_manual_bill: {e}")
        flash('An error occurred while processing manual bill generation', 'error')
        return redirect(url_for('bills'))

@app.route('/generate_multiple_bills', methods=['GET', 'POST'])
def generate_multiple_bills():
    """Show multiple bill generation page or process multiple bill generation"""
    try:
        if not hasattr(app, 'bills') or not app.bills:
            flash('No bills available. Please upload an Excel file first.', 'error')
            return redirect(url_for('index'))
        
        if request.method == 'POST':
            selected_centers = request.form.getlist('selected_centers')
            if not selected_centers:
                flash('Please select at least one center', 'error')
                return redirect(url_for('generate_multiple_bills'))
            
            # Filter bills for selected centers
            filtered_bills = [bill for bill in app.bills if bill['centre_name'] in selected_centers]
            
            if not filtered_bills:
                flash('No bills found for selected centers', 'error')
                return redirect(url_for('generate_multiple_bills'))
            
            app.bills = filtered_bills
            # --- Audit log for multiple bill generation ---
            user_id = session.get('user_id') if 'user_id' in session else None
            log = AuditLog(user_id=user_id, action='generate_multiple_bills', bill_id=None, details=f'Generated {len(filtered_bills)} bills for selected centers')
            db.session.add(log)
            db.session.commit()
            flash(f'Generated {len(filtered_bills)} bills for selected centers', 'success')
            return redirect(url_for('bills'))
        
        return render_template('multiple_bills.html', bills=app.bills, app=app)
    except Exception as e:
        logger.error(f"Error in generate_multiple_bills: {e}")
        flash('An error occurred while processing multiple bill generation', 'error')
        return redirect(url_for('bills'))

@app.route('/generate_hlm_bills', methods=['GET', 'POST'])
def generate_hlm_bills():
    """Two-step HLM bill generation: select center, then enter sharing per test type."""
    try:
        if not hasattr(app, 'df') or app.df.empty:
            flash('No data available. Please upload an Excel file first.', 'error')
            return redirect(url_for('index'))
        df = app.df.copy()
        # Step 1: Show HLM centers
        if request.method == 'GET' or (request.method == 'POST' and 'selected_center' not in request.form):
            hlm_df = df[df['MobileNumber'].astype(str).str.strip().str.upper() == 'HLM']
            centers = sorted(hlm_df['CENTER NAME'].dropna().unique())
            return render_template('hlm_bills.html', hlm_centers=centers, app=app)
        # Step 2: Center selected, show test types and sharing input
        selected_center = request.form.get('selected_center')
        if not selected_center:
            flash('Please select a center', 'error')
            return redirect(url_for('generate_hlm_bills'))
        center_df = df[(df['MobileNumber'].astype(str).str.strip().str.upper() == 'HLM') & (df['CENTER NAME'] == selected_center)]
        test_types = sorted(center_df['TEST TYPE'].dropna().unique())
        # If sharing percentages submitted, generate bill
        if 'submit_sharing' in request.form:
            sharing_map = {}
            for test_type in test_types:
                key = f"sharing_{test_type.replace(' ', '_').lower()}"
                val = request.form.get(key)
                try:
                    sharing_map[test_type] = int(val)
                except (TypeError, ValueError):
                    sharing_map[test_type] = 55.0
            # Generate bill for selected center
            test_items = []
            total_mrp = 0
            total_rate = 0
            total_sharing = 0
            for _, row in center_df.iterrows():
                mrp = safe_float_conversion(row['MRP'])
                test_type = str(row['TEST TYPE'])
                sharing_pct = sharing_map.get(test_type, 55.0)
                sharing_amount = mrp * (sharing_pct / 100)
                rate = mrp - sharing_amount
                test_item = {
                    'registered_date': safe_date_conversion(row['RegisteredDate']),
                    'visit_code': str(safe_int_conversion(row['PatientVisitCode'])),
                    'patient_name': str(row['PatientName']) if pd.notna(row['PatientName']) else 'N/A',
                    'test_name': str(row['TEST NAME']) if pd.notna(row['TEST NAME']) else 'N/A',
                    'test_type': test_type,
                    'mrp': mrp,
                    'rate': rate,
                    'sharing_amount': sharing_amount,
                    'sharing_percentage': sharing_pct
                }
                test_items.append(test_item)
                total_mrp += mrp
                total_rate += rate
                total_sharing += sharing_amount
            invoice_number = invoice_generator.generate(center_type='HLM', center_name=selected_center)
            bill = {
                'centre_name': selected_center,
                'test_items': test_items,
                'test_types': test_types,
                'total_mrp': total_mrp,
                'total_rate': total_rate,
                'total_sharing': total_sharing,
                'bill_date': datetime.now().strftime('%Y-%m-%d'),
                'bill_number': invoice_number,
                'center_type': 'HLM',
                'amount_in_words': amount_converter.convert(total_rate),
                'center_rows_data': center_df.to_dict('records')
            }
            app.bills = [bill]
              # 🔹 Excel Generation Call
            try:
                excel_path = generate_hlm_excel_from_template(bill, center_df.to_dict('records'))
                logger.info(f"HLM Excel generated: {excel_path}")
            except Exception as e:
                logger.error(f"Excel generation failed: {e}")

            # --- Audit log for HLM bill generation ---
            user_id = session.get('user_id') if 'user_id' in session else None
            log = AuditLog(user_id=user_id, action='generate_hlm_bill', bill_id=None, details=f'Generated HLM bill for {selected_center}')
            db.session.add(log)
            db.session.commit()
            flash(f'Generated HLM bill for {selected_center}', 'success')
            return redirect(url_for('bills'))

        return render_template('hlm_bills.html', selected_center=selected_center, test_types=test_types, app=app)
    except Exception as e:
        logger.error(f"Error in generate_hlm_bills: {e}")
        flash('An error occurred while processing HLM bill generation', 'error')
        return redirect(url_for('bills'))
          

@app.route('/generate_b2b_bills')
def generate_b2b_bills():
    """One-click B2B bill generation."""
    try:
        if not hasattr(app, 'df') or app.df.empty:
            flash('No data available. Please upload an Excel file first.', 'error')
            return redirect(url_for('index'))
        df = app.df.copy()
        b2b_df = df[df['MobileNumber'].astype(str).str.strip().str.upper() == 'B2B']
        bills = []
        month_str = datetime.now().strftime('%Y-%m')
        # Ensure B2B category exists
        category = BillCategory.query.filter_by(name='B2B').first()
        if not category:
            category = BillCategory(name='B2B', description='B2B Centers')
            db.session.add(category)
            db.session.commit()
        uploaded_file = UploadedFile.query.order_by(UploadedFile.id.desc()).first()
        for center_name, group in b2b_df.groupby('CENTER NAME'):
            if pd.isna(center_name) or center_name == '':
                continue
            test_items = []
            total_mrp = 0
            total_rate = 0
            total_sharing = 0
            for _, row in group.iterrows():
                mrp = safe_float_conversion(row['MRP'])
                rate = safe_float_conversion(row['CentreTestRate'])
                sharing_amount = mrp - rate
                test_item = {
                    'registered_date': safe_date_conversion(row['RegisteredDate']),
                    'visit_code': str(safe_int_conversion(row['PatientVisitCode'])),
                    'patient_name': str(row['PatientName']) if pd.notna(row['PatientName']) else 'N/A',
                    'test_name': str(row['TEST NAME']) if pd.notna(row['TEST NAME']) else 'N/A',
                    'test_type': str(row.get('TEST TYPE', 'Other')).strip(),
                    'mrp': mrp,
                    'rate': rate,
                    'sharing_amount': sharing_amount
                }
                test_items.append(test_item)
                total_mrp += mrp
                total_rate += rate
                total_sharing += sharing_amount
            invoice_number = invoice_generator.generate(center_type='B2B', center_name=str(center_name))
            bill_data = {
                'centre_name': str(center_name),
                'test_items': test_items,
                'total_mrp': total_mrp,
                'total_rate': total_rate,
                'total_sharing': total_sharing,
                'bill_date': datetime.now().strftime('%Y-%m-%d'),
                'bill_number': invoice_number,
                'center_type': 'B2B',
                'amount_in_words': amount_converter.convert(total_rate)
            }
            # Store bill in DB
            bill_db = Bill(
                bill_number=invoice_number,
                center_name=str(center_name),
                month=month_str,
                category_id=category.id,
                uploaded_file_id=uploaded_file.id if uploaded_file else None,
                bill_data=bill_data
            )
            db.session.add(bill_db)
            bills.append(bill_data)
        db.session.commit()
        if not bills:
            flash('No B2B bills found in the uploaded data', 'error')
            return redirect(url_for('bills'))
        app.bills = bills
        # --- Audit log for B2B bill generation ---
        user_id = session.get('user_id') if 'user_id' in session else None
        log = AuditLog(user_id=user_id, action='generate_b2b_bills', bill_id=None, details=f'Generated {len(bills)} B2B bills')
        db.session.add(log)
        db.session.commit()
        flash(f'Generated {len(bills)} B2B bills', 'success')
        return redirect(url_for('bills'))
    except Exception as e:
        logger.error(f"Error in generate_b2b_bills: {e}")
        flash('An error occurred while processing B2B bill generation', 'error')
        return redirect(url_for('bills'))


# --- New: Bills listing with DB filtering ---
@app.route('/bills')
def bills():
    try:
        # Get filter params
        month = request.args.get('month')
        category_name = request.args.get('category')
        query = Bill.query
        if month:
            query = query.filter(Bill.month == month)
        if category_name:
            category = BillCategory.query.filter_by(name=category_name).first()
            if category:
                query = query.filter(Bill.category_id == category.id)
        bills_db = query.order_by(Bill.created_at.desc()).all()
        bills = [b.bill_data for b in bills_db]
        if not bills:
            flash('No bills available for the selected filter.', 'error')
            return redirect(url_for('index'))
        # Calculate totals
        total_tests = sum(len(b['test_items']) for b in bills)
        total_mrp = sum(b.get('total_mrp', 0) for b in bills)
        total_rate = sum(b.get('total_rate', 0) for b in bills)
        total_sharing = sum(b.get('total_sharing', 0) for b in bills)
        return render_template('bills.html',
                             bills=bills,
                             total_tests=total_tests,
                             total_mrp=total_mrp,
                             total_rate=total_rate,
                             total_sharing=total_sharing,
                             app=app)
    except Exception as e:
        logger.error(f"Error in bills route: {e}")
        flash('An error occurred while loading bills', 'error')
        return redirect(url_for('index'))

# --- New: API endpoint for bills by month/category ---
@app.route('/api/bills/filter')
def api_bills_filter():
    try:
        month = request.args.get('month')
        category_name = request.args.get('category')
        query = Bill.query
        if month:
            query = query.filter(Bill.month == month)
        if category_name:
            category = BillCategory.query.filter_by(name=category_name).first()
            if category:
                query = query.filter(Bill.category_id == category.id)
        bills_db = query.order_by(Bill.created_at.desc()).all()
        bills = [b.bill_data for b in bills_db]
        return jsonify({'bills': bills, 'count': len(bills)})
    except Exception as e:
        logger.error(f"Error in api_bills_filter: {e}")
        return jsonify({'error': 'Internal server error'}), 500

@app.route('/bill/<int:bill_index>')
def view_bill(bill_index):
    try:
        if not hasattr(app, 'bills') or not app.bills:
            flash('No bills available', 'error')
            return redirect(url_for('bills'))
        if bill_index < 0 or bill_index >= len(app.bills):
            flash('Bill not found', 'error')
            return redirect(url_for('bills'))
        bill = app.bills[bill_index]
        # Get payment info from DB if available
        bill_obj = Bill.query.filter_by(bill_number=bill['bill_number']).first()
        payment_mode = bill_obj.payment_mode if bill_obj else None
        payment_details = bill_obj.payment_details if bill_obj else None
        return render_template('bill_detail.html', bill=bill, bill_index=bill_index, app=app, payment_mode=payment_mode, payment_details=payment_details)
    except Exception as e:
        logger.error(f"Error in view_bill: {e}")
        flash('An error occurred while viewing the bill', 'error')
        return redirect(url_for('bills'))

@app.route('/download_bill/<int:bill_index>')
def download_bill(bill_index):
    try:
        if not hasattr(app, 'bills') or not app.bills:
            flash('No bills available', 'error')
            return redirect(url_for('bills'))
        if bill_index < 0 or bill_index >= len(app.bills):
            flash('Bill not found', 'error')
            return redirect(url_for('bills'))

        bill = app.bills[bill_index]
        fmt = request.args.get('format', 'html').lower()
        safe_center_name = str(bill['centre_name']).replace(" ", "_").replace("/", "_")

        if fmt == 'excel':
            if bill.get("center_type") == "HLM":
                excel_path = generate_hlm_excel_from_template(bill, bill.get('center_rows_data', []))
                return send_file(
                    excel_path,
                    mimetype="application/vnd.ms-excel",
                    as_attachment=True,
                    download_name=f"{safe_center_name}.xlsm"
                )
            else:
                buffer = BytesIO()
                df = pd.DataFrame(bill['test_items'])
                df.to_excel(buffer, index=False)
                buffer.seek(0)
                filename = f"{safe_center_name}.xlsx"
                return send_file(
                    buffer,
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    as_attachment=True,
                    download_name=filename
                )

        elif fmt == 'pdf':
            html_content = render_template('bill_pdf.html', bill=bill)
            pdf_buffer = BytesIO()
            try:
                import pdfkit
                pdf = pdfkit.from_string(html_content, False)
                pdf_buffer.write(pdf)
                pdf_buffer.seek(0)
            except Exception:
                from xhtml2pdf import pisa
                pisa_status = pisa.CreatePDF(html_content, dest=pdf_buffer)
                if pisa_status.err:
                    raise Exception("xhtml2pdf generation failed")
                pdf_buffer.seek(0)

            filename = f"{safe_center_name}.pdf"
            return send_file(pdf_buffer, mimetype='application/pdf',
                             as_attachment=True, download_name=filename)

        else:
            html_content = render_template('bill_pdf.html', bill=bill)
            with tempfile.NamedTemporaryFile(mode='w', suffix='.html', delete=False, encoding='utf-8') as f:
                f.write(html_content)
                temp_path = f.name
            return send_file(temp_path, as_attachment=True, download_name=f"{bill['bill_number']}.html")

    except Exception as e:
        logger.error(f"Error in download_bill: {e}")
        flash('An error occurred while downloading the bill', 'error')
        return redirect(url_for('view_bill', bill_index=bill_index))

@app.route('/download_all_bills')
def download_all_bills():
    try:
        if not hasattr(app, 'bills') or not app.bills:
            flash('No bills available', 'error')
            return redirect(url_for('index'))
        
        # Create a ZIP file containing all bills
        memory_file = BytesIO()
        
        with zipfile.ZipFile(memory_file, 'w') as zf:
            for i, bill in enumerate(app.bills):
                try:
                    html_content = render_template('bill_pdf.html', bill=bill)
                    zf.writestr(f"{bill['bill_number']}.html", html_content)
                except Exception as e:
                    logger.error(f"Error processing bill {i}: {e}")
                    continue
        
        memory_file.seek(0)
        
        return send_file(
            memory_file,
            mimetype='application/zip',
            as_attachment=True,
            download_name=f"all_bills_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
        )
    except Exception as e:
        logger.error(f"Error in download_all_bills: {e}")
        flash('An error occurred while downloading all bills', 'error')
        return redirect(url_for('bills'))

@app.route('/download_all_excel')
def download_all_excel():
    try:
        if not hasattr(app, 'bills') or not app.bills:
            flash('No bills available', 'error')
            return redirect(url_for('bills'))

        memory_file = BytesIO()
        with zipfile.ZipFile(memory_file, 'w') as zf:
            for bill in app.bills:
                safe_center_name = str(bill['centre_name']).replace(" ", "_").replace("/", "_")
                if bill.get("center_type") == "HLM":
                    excel_path = generate_hlm_excel_from_template(bill, bill.get('center_rows_data', []))
                    zf.write(excel_path, arcname=f"{safe_center_name}.xlsm")
                else:
                    buffer = BytesIO()
                    df = pd.DataFrame(bill['test_items'])
                    df.to_excel(buffer, index=False)
                    buffer.seek(0)
                    zf.writestr(f"{safe_center_name}.xlsx", buffer.read())

        memory_file.seek(0)
        return send_file(
            memory_file,
            mimetype='application/zip',
            as_attachment=True,
            download_name=f"all_bills_excel_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
        )
    except Exception as e:
        logger.error(f"Error in download_all_excel: {e}")
        flash('An error occurred while downloading all Excel bills', 'error')
        return redirect(url_for('bills'))

@app.route('/download_all_pdf')
def download_all_pdf():
    try:
        if not hasattr(app, 'bills') or not app.bills:
            flash('No bills available', 'error')
            return redirect(url_for('bills'))

        memory_file = BytesIO()
        with zipfile.ZipFile(memory_file, 'w') as zf:
            for bill in app.bills:
                safe_center_name = str(bill['centre_name']).replace(" ", "_").replace("/", "_")
                html_content = render_template('bill_pdf.html', bill=bill)
                pdf_buffer = BytesIO()
                try:
                    import pdfkit
                    pdf = pdfkit.from_string(html_content, False)
                    pdf_buffer.write(pdf)
                except Exception:
                    from xhtml2pdf import pisa
                    pisa_status = pisa.CreatePDF(html_content, dest=pdf_buffer)
                    if pisa_status.err:
                        continue
                pdf_buffer.seek(0)
                zf.writestr(f"{safe_center_name}.pdf", pdf_buffer.read())

        memory_file.seek(0)
        return send_file(
            memory_file,
            mimetype='application/zip',
            as_attachment=True,
            download_name=f"all_bills_pdf_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
        )
    except Exception as e:
        logger.error(f"Error in download_all_pdf: {e}")
        flash('An error occurred while downloading all PDF bills', 'error')
        return redirect(url_for('bills'))

@app.route('/api/bills')
def api_bills():
    try:
        if not hasattr(app, 'bills') or not app.bills:
            return jsonify({'error': 'No bills available'}), 404
        
        return jsonify(app.bills)
    except Exception as e:
        logger.error(f"Error in api_bills: {e}")
        return jsonify({'error': 'Internal server error'}), 500

@app.route('/ai_assistance', methods=['GET', 'POST'])
def ai_assistance():
    """AI assistance page for user queries"""
    if request.method == 'POST':
        user_query = request.form.get('user_query', '')
        if user_query:
            try:
                ai_response = ai_integration.get_ai_response(user_query)
                return jsonify({'response': ai_response})
            except Exception as e:
                logger.error(f"AI assistance error: {e}")
                return jsonify({'response': f"AI assistance temporarily unavailable: {str(e)}"})
    
    return render_template('ai_assistance.html', app=app)

# Error handlers
@app.errorhandler(404)
def not_found_error(error):
    return render_template('404.html'), 404

@app.errorhandler(500)
def internal_error(error):
    logger.error(f"Internal server error: {error}")
    return render_template('500.html'), 500

# if __name__ == '__main__':
#    app.run(debug=True, host='0.0.0.0', port=5000) 